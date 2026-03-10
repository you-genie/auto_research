"""
참고문헌 XLSX 파일 생성 스크립트
파이썬 개발자를 위한 클린코드 가이드 — Python 3.12/3.13 신기능 중심
"""

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

REFERENCES = [
    {
        "번호": 1,
        "제목": "What's New In Python 3.12",
        "저자/출처": "Python Software Foundation",
        "URL": "https://docs.python.org/3/whatsnew/3.12.html",
        "발행일": "2023-10-02",
        "요약 (한국어)": "Python 3.12 공식 변경사항 문서. PEP 701 f-string 개선, PEP 695 타입 파라미터 문법, PEP 698 @override 데코레이터, PEP 684 Per-Interpreter GIL, 향상된 에러 메시지 등 주요 신기능 설명.",
        "관련 섹션": "Python 3.12 신기능",
    },
    {
        "번호": 2,
        "제목": "What's New In Python 3.13",
        "저자/출처": "Python Software Foundation",
        "URL": "https://docs.python.org/3/whatsnew/3.13.html",
        "발행일": "2024-10-07",
        "요약 (한국어)": "Python 3.13 공식 변경사항 문서. 개선된 REPL, Free-threaded 모드(PEP 703), JIT 컴파일러(PEP 744), TypeIs(PEP 742), ReadOnly TypedDict(PEP 705), locals() 의미론 변경(PEP 667) 등 포함.",
        "관련 섹션": "Python 3.13 신기능",
    },
    {
        "번호": 3,
        "제목": "PEP 701 — Syntactic Formalization of f-strings",
        "저자/출처": "Pablo Galindo Salgado (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0701/",
        "발행일": "2022-11-19",
        "요약 (한국어)": "f-string의 문법적 공식화를 다룬 PEP. 따옴표 재사용, 백슬래시 허용, 여러 줄 표현식 및 주석 지원 등 f-string의 오랜 제약을 제거. 토크나이저 성능 64% 향상.",
        "관련 섹션": "Python 3.12 신기능 - f-string",
    },
    {
        "번호": 4,
        "제목": "PEP 695 — Type Parameter Syntax",
        "저자/출처": "Eric Traut (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0695/",
        "발행일": "2022-06-16",
        "요약 (한국어)": "제네릭 클래스/함수의 새로운 타입 파라미터 문법 도입. def func[T](...), class Cls[T], type X = ... 문법으로 TypeVar 명시적 임포트 불필요. 더 간결하고 읽기 쉬운 제네릭 코드 작성 가능.",
        "관련 섹션": "Python 3.12 신기능 - 타입 파라미터",
    },
    {
        "번호": 5,
        "제목": "PEP 698 — Override Decorator for Static Typing",
        "저자/출처": "Steven Troxler (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0698/",
        "발행일": "2022-09-05",
        "요약 (한국어)": "@override 데코레이터 도입. 서브클래스에서 부모 메서드를 재정의할 때 명시적 의도 표현. 타입 체커가 메서드 이름 오타나 시그니처 불일치를 조기에 감지.",
        "관련 섹션": "Python 3.12 신기능 - @override",
    },
    {
        "번호": 6,
        "제목": "PEP 684 — A Per-Interpreter GIL",
        "저자/출처": "Eric Snow (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0684/",
        "발행일": "2021-11-16",
        "요약 (한국어)": "서브인터프리터마다 독립적인 GIL 허용. 아직 C-API 수준으로 제한적이지만 Python 3.13 Free-Threading의 기반. AI 추론 서버에서 병렬 처리를 위한 초석.",
        "관련 섹션": "Python 3.12 신기능 - Per-Interpreter GIL",
    },
    {
        "번호": 7,
        "제목": "PEP 703 — Making the Global Interpreter Lock Optional",
        "저자/출처": "Sam Gross (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0703/",
        "발행일": "2022-06-08",
        "요약 (한국어)": "GIL을 선택적으로 비활성화하는 Free-Threaded CPython 실험적 빌드. PYTHON_GIL 환경 변수로 제어. 단일 스레드 성능 약 30~40% 저하, ML 라이브러리 별도 호환 빌드 필요. 프로덕션 미권장.",
        "관련 섹션": "Python 3.13 신기능 - Free-Threading",
    },
    {
        "번호": 8,
        "제목": "PEP 744 — JIT Compilation",
        "저자/출처": "Brandt Bucher (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0744/",
        "발행일": "2023-11-05",
        "요약 (한국어)": "Copy-and-patch 기법 기반의 실험적 JIT 컴파일러. 기본 비활성화이며 PYTHON_JIT 환경 변수로 제어. 현재 몇 퍼센트 수준 성능 향상. AI/ML 워크로드에는 Numba 등이 더 효과적.",
        "관련 섹션": "Python 3.13 신기능 - JIT",
    },
    {
        "번호": 9,
        "제목": "PEP 742 — Narrowing types with TypeIs",
        "저자/출처": "Randolph Voorhies et al. (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0742/",
        "발행일": "2024-01-04",
        "요약 (한국어)": "TypeGuard의 한계를 보완하는 TypeIs 타입 가드 도입. 더 정확한 타입 내로잉, 특히 제네릭 타입과 유니온 타입에서 개선된 동작. Python 3.13에서 추가.",
        "관련 섹션": "Python 3.13 신기능 - 타입 시스템",
    },
    {
        "번호": 10,
        "제목": "PEP 667 — Consistent views of namespaces",
        "저자/출처": "Mark Shannon (Python Enhancement Proposals)",
        "URL": "https://peps.python.org/pep-0667/",
        "발행일": "2021-09-20",
        "요약 (한국어)": "함수 내 locals()가 독립적인 스냅샷을 반환하도록 의미론을 명확화. 디버거와 메타프로그래밍 코드에서 예측 가능한 동작 보장. FrameType.f_locals는 쓰기 가능 프록시로 제공.",
        "관련 섹션": "Python 3.13 신기능 - locals()",
    },
    {
        "번호": 11,
        "제목": "Python 3.13: Cool New Features for You to Try",
        "저자/출처": "Real Python",
        "URL": "https://realpython.com/python313-new-features/",
        "발행일": "2024-10-01",
        "요약 (한국어)": "Python 3.13 신기능 실용적 소개. 새 REPL 개선사항, 컬러 트레이스백, 키워드 인수 오타 제안, 타입 파라미터 기본값, TypeIs 등 개발자 경험 향상 기능 코드 예시와 함께 설명.",
        "관련 섹션": "Python 3.13 신기능",
    },
    {
        "번호": 12,
        "제목": "Python 3.13: Free Threading and a JIT Compiler",
        "저자/출처": "Real Python",
        "URL": "https://realpython.com/python313-free-threading-jit/",
        "발행일": "2024-10-07",
        "요약 (한국어)": "Python 3.13의 Free-Threaded 모드와 JIT 컴파일러 심층 분석. 활성화 방법, 성능 벤치마크, AI/ML 워크로드에서의 현실적 한계와 주의사항 포함.",
        "관련 섹션": "Python 3.13 신기능 - Free-Threading & JIT",
    },
    {
        "번호": 13,
        "제목": "FastAPI Best Practices and Conventions",
        "저자/출처": "Zhanymkanov (GitHub)",
        "URL": "https://github.com/zhanymkanov/fastapi-best-practices",
        "발행일": "2023-01-01",
        "요약 (한국어)": "FastAPI 실전 베스트 프랙티스 모음. 도메인 중심 프로젝트 구조, 의존성 주입 패턴, 비동기 라우터 설계, Pydantic 적극 활용, 에러 핸들링, 응답 모델 설계 등 실무 패턴 정리.",
        "관련 섹션": "AI 백엔드 패턴",
    },
    {
        "번호": 14,
        "제목": "FastAPI Settings and Environment Variables",
        "저자/출처": "FastAPI 공식 문서",
        "URL": "https://fastapi.tiangolo.com/advanced/settings/",
        "발행일": "2024-01-01",
        "요약 (한국어)": "FastAPI에서 pydantic-settings를 활용한 환경 변수 관리 공식 가이드. BaseSettings 사용법, .env 파일 통합, @lru_cache를 활용한 설정 싱글턴 패턴, 테스트에서 설정 오버라이드 방법.",
        "관련 섹션": "AI 백엔드 패턴 - 설정 관리",
    },
    {
        "번호": 15,
        "제목": "Pydantic Settings — Settings Management",
        "저자/출처": "Pydantic 공식 문서",
        "URL": "https://docs.pydantic.dev/latest/concepts/pydantic_settings/",
        "발행일": "2024-06-01",
        "요약 (한국어)": "pydantic-settings 공식 문서. BaseSettings 클래스, SecretStr, 환경 변수 파싱, 중첩 설정, env_prefix 등 타입-안전한 설정 관리를 위한 모든 기능 설명.",
        "관련 섹션": "AI 백엔드 패턴 - 설정 관리",
    },
    {
        "번호": 16,
        "제목": "Retry Logic with Tenacity — Instructor",
        "저자/출처": "Instructor (python.useinstructor.com)",
        "URL": "https://python.useinstructor.com/concepts/retrying/",
        "발행일": "2024-03-01",
        "요약 (한국어)": "Tenacity 라이브러리를 활용한 LLM API 재시도 로직 구현. @retry 데코레이터, 지수 백오프 wait 설정, RateLimitError/APITimeoutError 처리, 검증 실패 시 LLM에 힌트 재전달 패턴.",
        "관련 섹션": "AI 백엔드 패턴 - 에러 핸들링",
    },
    {
        "번호": 17,
        "제목": "Asynchronous LLM API Calls in Python: A Comprehensive Guide",
        "저자/출처": "Unite.AI",
        "URL": "https://www.unite.ai/asynchronous-llm-api-calls-in-python-a-comprehensive-guide/",
        "발행일": "2024-09-01",
        "요약 (한국어)": "Python에서 비동기 LLM API 호출 종합 가이드. asyncio 기반 동시 요청 처리, 배치 처리, 타임아웃 설정, 오류 핸들링, 속도 제한 대응 등 실전 패턴.",
        "관련 섹션": "AI 백엔드 패턴 - 비동기 패턴",
    },
    {
        "번호": 18,
        "제목": "Error Handling Best Practices for Production LLM Applications",
        "저자/출처": "Markaicode",
        "URL": "https://markaicode.com/llm-error-handling-production-guide/",
        "발행일": "2024-08-01",
        "요약 (한국어)": "프로덕션 LLM 애플리케이션의 에러 핸들링 완전 가이드. 예외 계층 설계, 재시도 전략, 폴백 메커니즘, 로깅, 카스케이드 장애 방지 패턴.",
        "관련 섹션": "AI 백엔드 패턴 - 에러 핸들링",
    },
    {
        "번호": 19,
        "제목": "Streamlit vs Gradio in 2025: Comparing AI-App Frameworks",
        "저자/출처": "Squadbase Blog",
        "URL": "https://www.squadbase.dev/en/blog/streamlit-vs-gradio-in-2025-a-framework-comparison-for-ai-apps",
        "발행일": "2025-01-01",
        "요약 (한국어)": "2025년 Streamlit vs Gradio 비교 분석. 상태 관리 방식 차이, 컴포넌트 분리 패턴, 확장성 한계, 실제 AI 앱 구축 시 선택 기준 제시.",
        "관련 섹션": "AI 프론트엔드 패턴",
    },
    {
        "번호": 20,
        "제목": "State In Blocks — Gradio 공식 문서",
        "저자/출처": "Gradio",
        "URL": "https://www.gradio.app/4.44.1/guides/state-in-blocks",
        "발행일": "2024-09-01",
        "요약 (한국어)": "Gradio Blocks API에서의 상태 관리 공식 가이드. 전역 상태와 사용자별 세션 상태 차이, gr.State 컴포넌트 사용법, 대화 히스토리 관리 패턴.",
        "관련 섹션": "AI 프론트엔드 패턴 - Gradio",
    },
    {
        "번호": 21,
        "제목": "4 Best Prompt Management Systems for LLM Developers in 2025",
        "저자/출처": "Mirascope",
        "URL": "https://mirascope.com/blog/prompt-management-system",
        "발행일": "2025-01-01",
        "요약 (한국어)": "LLM 개발자를 위한 프롬프트 관리 시스템 비교. Langfuse, PromptLayer, Banks 등 도구 비교. 프롬프트를 코드처럼 버전 관리하고 PR 리뷰, CI 테스트하는 방법론 소개.",
        "관련 섹션": "LLM 공통 패턴 - 프롬프트 관리",
    },
    {
        "번호": 22,
        "제목": "Best Prompt Versioning Tools for LLM Optimization (2025)",
        "저자/출처": "PromptLayer Blog",
        "URL": "https://blog.promptlayer.com/5-best-tools-for-prompt-versioning/",
        "발행일": "2025-01-15",
        "요약 (한국어)": "프롬프트 버전 관리 도구 Top 5 분석. 각 도구의 특징, 프롬프트 버전 관리의 필요성, 팀 협업에서의 프롬프트 관리 모범 사례.",
        "관련 섹션": "LLM 공통 패턴 - 프롬프트 관리",
    },
    {
        "번호": 23,
        "제목": "Production-Ready RAG Pipelines with Haystack and LangChain",
        "저자/출처": "DigitalOcean",
        "URL": "https://www.digitalocean.com/community/tutorials/production-ready-rag-pipelines-haystack-langchain",
        "발행일": "2024-11-01",
        "요약 (한국어)": "Haystack과 LangChain으로 프로덕션 수준 RAG 파이프라인 구축 튜토리얼. 문서 처리, 임베딩 생성, 벡터 DB 통합, 검색, 컨텍스트 조립, 생성 단계의 모듈형 설계 패턴.",
        "관련 섹션": "LLM 공통 패턴 - RAG 파이프라인",
    },
    {
        "번호": 24,
        "제목": "Complete Guide to Building a Robust RAG Pipeline 2025",
        "저자/출처": "Dhiwise",
        "URL": "https://www.dhiwise.com/post/build-rag-pipeline-guide",
        "발행일": "2025-02-01",
        "요약 (한국어)": "2025년 기준 RAG 파이프라인 구축 완전 가이드. 청킹 전략, 벡터 검색 최적화, 컨텍스트 어셈블리, 생성 품질 개선, 평가 방법, 확장성 고려사항.",
        "관련 섹션": "LLM 공통 패턴 - RAG 파이프라인",
    },
    {
        "번호": 25,
        "제목": "Ruff — An Extremely Fast Python Linter and Formatter",
        "저자/출처": "Astral (GitHub)",
        "URL": "https://github.com/astral-sh/ruff",
        "발행일": "2022-08-01",
        "요약 (한국어)": "Rust로 작성된 초고속 Python 린터 및 포매터. Black, Flake8, isort, pyupgrade 등 다수 도구를 단일 도구로 대체. pyproject.toml 통합 설정, pre-commit 훅 지원.",
        "관련 섹션": "도구 추천 - Ruff",
    },
    {
        "번호": 26,
        "제목": "Ruff FAQ",
        "저자/출처": "Astral",
        "URL": "https://docs.astral.sh/ruff/faq/",
        "발행일": "2024-01-01",
        "요약 (한국어)": "Ruff 자주 묻는 질문. mypy/pyright와의 보완적 관계, Black과의 차이, 규칙 선택 가이드, 기존 도구에서 마이그레이션 방법.",
        "관련 섹션": "도구 추천 - Ruff",
    },
    {
        "번호": 27,
        "제목": "Effortless Code Quality: The Ultimate Pre-Commit Hooks Guide for 2025",
        "저자/출처": "Gatlen Culp (Medium)",
        "URL": "https://gatlenculp.medium.com/effortless-code-quality-the-ultimate-pre-commit-hooks-guide-for-2025-57ca501d9835",
        "발행일": "2025-01-10",
        "요약 (한국어)": "2025년 Python 프로젝트를 위한 pre-commit 훅 완전 가이드. Ruff, pyright, detect-private-key 등 AI 개발에 필수적인 훅 설정 방법과 .pre-commit-config.yaml 예시.",
        "관련 섹션": "도구 추천 - pre-commit",
    },
    {
        "번호": 28,
        "제목": "Unveiling Python 3.12: A Leap Forward",
        "저자/출처": "JetBrains PyCharm Blog",
        "URL": "https://blog.jetbrains.com/pycharm/2023/11/python-3-12/",
        "발행일": "2023-11-01",
        "요약 (한국어)": "PyCharm 관점에서 Python 3.12 신기능 소개. 타입 파라미터 문법(PEP 695), f-string 개선(PEP 701)이 IDE 자동완성과 타입 추론에 미치는 영향 분석.",
        "관련 섹션": "Python 3.12 신기능",
    },
    {
        "번호": 29,
        "제목": "Pydantic AI — FastAPI for GenAI App Development",
        "저자/출처": "Pydantic AI",
        "URL": "https://ai.pydantic.dev/",
        "발행일": "2024-11-01",
        "요약 (한국어)": "Pydantic 팀이 만든 GenAI 에이전트 개발 프레임워크. FastAPI와 동일한 설계 철학 적용. FastAPI와의 통합, 타입-안전한 에이전트 도구 정의, 구조화된 응답 처리.",
        "관련 섹션": "AI 백엔드 패턴",
    },
    {
        "번호": 30,
        "제목": "Fallbacks and Retries — LiteLLM",
        "저자/출처": "DeepWiki / BerriAI",
        "URL": "https://deepwiki.com/BerriAI/litellm/7.1-fallbacks-and-retries",
        "발행일": "2024-06-01",
        "요약 (한국어)": "LiteLLM의 폴백 및 재시도 전략 문서. 기본 폴백, 컨텍스트 윈도우 폴백, 콘텐츠 정책 폴백 등 세 가지 폴백 전략. 비동기 폴백 처리 async_completion_with_fallbacks.",
        "관련 섹션": "AI 백엔드 패턴 - 에러 핸들링",
    },
]


def create_xlsx(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 스타일 정의
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Malgun Gothic", size=10)
    link_font = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 행
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 45, 25, 50, 12, 60, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, ref in enumerate(REFERENCES, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약 (한국어)"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            ws.row_dimensions[row_idx].height = 60

            if col_idx == 1:  # 번호
                cell.alignment = center_align
                cell.font = Font(name="Malgun Gothic", size=10, bold=True)
            elif col_idx == 4:  # URL — 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = link_font
                cell.alignment = left_align
            elif col_idx == 5:  # 발행일
                cell.alignment = center_align
                cell.font = body_font
            else:
                cell.alignment = left_align
                cell.font = body_font

    # 틀 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "1F4E79"

    wb.save(output_path)
    print(f"XLSX 파일 생성 완료: {output_path}")
    print(f"총 참고문헌 수: {len(REFERENCES)}개")


if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "python-cleancode-guide-references.xlsx")
    create_xlsx(output_path)
