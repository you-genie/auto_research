"""
참고문헌 XLSX 파일 생성 스크립트
=================================
실행: python create_references.py
출력: python-chatbot-tools-references.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ─── 참고문헌 데이터 ────────────────────────────────────────────────
REFERENCES = [
    {
        "번호": 1,
        "제목": "Build a basic LLM chat app - Streamlit Docs",
        "저자/출처": "Streamlit 공식 문서",
        "URL": "https://docs.streamlit.io/develop/tutorials/chat-and-llm-apps/build-conversational-apps",
        "발행일": "2024",
        "요약": "Streamlit의 st.chat_message, st.chat_input, st.write_stream을 사용하여 에코봇부터 GPT 스타일 챗봇까지 단계적으로 구현하는 공식 튜토리얼",
        "관련 섹션": "Streamlit 챗봇 구현",
    },
    {
        "번호": 2,
        "제목": "How to build an LLM-powered ChatBot with Streamlit",
        "저자/출처": "Streamlit Blog",
        "URL": "https://blog.streamlit.io/how-to-build-an-llm-powered-chatbot-with-streamlit/",
        "발행일": "2024",
        "요약": "Streamlit으로 LLM 챗봇을 구축하는 공식 블로그 포스트. session_state, 스트리밍, OpenAI 통합 방법 설명",
        "관련 섹션": "Streamlit 챗봇 구현",
    },
    {
        "번호": 3,
        "제목": "RAG and Streamlit Chatbot: Chat with Documents Using LLM",
        "저자/출처": "Analytics Vidhya",
        "URL": "https://www.analyticsvidhya.com/blog/2024/04/rag-and-streamlit-chatbot-chat-with-documents-using-llm/",
        "발행일": "2024-04",
        "요약": "Streamlit + RAG(검색 증강 생성) 조합으로 문서 기반 챗봇을 만드는 방법. 파일 업로드, 임베딩, 검색 과정 포함",
        "관련 섹션": "Streamlit 배포 및 활용",
    },
    {
        "번호": 4,
        "제목": "Creating a Chatbot Fast - Gradio Guides",
        "저자/출처": "Gradio 공식 문서",
        "URL": "https://www.gradio.app/guides/creating-a-chatbot-fast",
        "발행일": "2024",
        "요약": "gr.ChatInterface를 사용하여 최소한의 코드로 완전한 챗봇을 만드는 공식 가이드. 스트리밍, 멀티모달, 커스터마이징 옵션 설명",
        "관련 섹션": "Gradio 챗봇 구현",
    },
    {
        "번호": 5,
        "제목": "Chatbot - Gradio Docs",
        "저자/출처": "Gradio 공식 문서",
        "URL": "https://www.gradio.app/docs/gradio/chatbot",
        "발행일": "2024",
        "요약": "gr.Chatbot 컴포넌트의 완전한 API 레퍼런스. 멀티미디어 메시지, 인용구, Chain-of-Thought 표시 기능 설명",
        "관련 섹션": "Gradio 고급 기능",
    },
    {
        "번호": 6,
        "제목": "Gradio and LLM Agents",
        "저자/출처": "Gradio 공식 문서",
        "URL": "https://www.gradio.app/guides/gradio-and-llm-agents",
        "발행일": "2024",
        "요약": "Gradio로 LLM 에이전트 UI를 구현하는 방법. Tool Calling, 중간 사고 과정(CoT), 반응형 버튼 시각화",
        "관련 섹션": "Gradio LLM 에이전트",
    },
    {
        "번호": 7,
        "제목": "ChatInterface Examples - Gradio",
        "저자/출처": "Gradio 공식 문서",
        "URL": "https://www.gradio.app/guides/chatinterface-examples",
        "발행일": "2024",
        "요약": "gr.ChatInterface의 다양한 사용 예시. 스트리밍, 멀티모달, 추가 입력 컴포넌트 통합 방법",
        "관련 섹션": "Gradio 코드 예시",
    },
    {
        "번호": 8,
        "제목": "GitHub - Chainlit/chainlit: Build Conversational AI in minutes",
        "저자/출처": "Chainlit GitHub",
        "URL": "https://github.com/Chainlit/chainlit",
        "발행일": "2024-2025",
        "요약": "Chainlit 공식 GitHub 저장소. LLM 챗봇 전용 오픈소스 프레임워크. 데코레이터 기반 아키텍처, 스트리밍, 인증, LangChain 통합 지원",
        "관련 섹션": "Chainlit 개요",
    },
    {
        "번호": 9,
        "제목": "It's 2025 — Start using Chainlit for your LLM Apps",
        "저자/출처": "Medium - MITB For All",
        "URL": "https://medium.com/mitb-for-all/its-2025-start-using-chainlit-for-your-llm-apps-558db1a46315",
        "발행일": "2025",
        "요약": "2025년 Chainlit 사용을 권장하는 이유. 프로덕션 수준 채팅 UI, 관찰 가능성, 상태 관리 기능 소개",
        "관련 섹션": "Chainlit 특징 및 장점",
    },
    {
        "번호": 10,
        "제목": "Building a Chatbot Application with Chainlit and LangChain",
        "저자/출처": "Medium - Tahreem Rasul",
        "URL": "https://medium.com/@tahreemrasul/building-a-chatbot-application-with-chainlit-and-langchain-3e86da0099a6",
        "발행일": "2024",
        "요약": "Chainlit + LangChain 조합으로 RAG 챗봇을 구현하는 단계별 튜토리얼",
        "관련 섹션": "Chainlit LangChain 통합",
    },
    {
        "번호": 11,
        "제목": "Reflex Chatapp Tutorial",
        "저자/출처": "Reflex 공식 문서",
        "URL": "https://reflex.dev/docs/getting-started/chatapp-tutorial/",
        "발행일": "2024-2025",
        "요약": "Reflex로 GPT 스타일 챗앱을 만드는 공식 튜토리얼. State 관리, 컴포넌트 구성, OpenAI 통합, 배포 방법 포함",
        "관련 섹션": "Reflex 챗봇 구현",
    },
    {
        "번호": 12,
        "제목": "Build a ChatGPT-esque Web App in Pure Python using Reflex",
        "저자/출처": "Towards Data Science",
        "URL": "https://towardsdatascience.com/build-a-chatgpt-esque-web-app-in-pure-python-using-reflex-bdc585038110/",
        "발행일": "2024",
        "요약": "파이썬만으로 ChatGPT와 유사한 웹앱을 Reflex로 구현하는 방법. React 없이 풀스택 앱 개발",
        "관련 섹션": "Reflex 아키텍처",
    },
    {
        "번호": 13,
        "제목": "GitHub - reflex-dev/reflex-chat: A ChatGPT clone built in Reflex",
        "저자/출처": "Reflex GitHub",
        "URL": "https://github.com/reflex-dev/reflex-chat",
        "발행일": "2024-2025",
        "요약": "Reflex로 만든 ChatGPT 클론 공식 예제 저장소. 실제 작동하는 완전한 코드 포함",
        "관련 섹션": "Reflex 코드 예시",
    },
    {
        "번호": 14,
        "제목": "Panel ChatInterface Docs",
        "저자/출처": "HoloViz Panel 공식 문서",
        "URL": "https://panel.holoviz.org/reference/chat/ChatInterface.html",
        "발행일": "2024-2025",
        "요약": "Panel ChatInterface 컴포넌트 API 레퍼런스. 파라미터, 이벤트 핸들러, 커스터마이징 옵션 설명",
        "관련 섹션": "Panel 챗봇 구현",
    },
    {
        "번호": 15,
        "제목": "GitHub - holoviz-topics/panel-chat-examples",
        "저자/출처": "HoloViz GitHub",
        "URL": "https://github.com/holoviz-topics/panel-chat-examples",
        "발행일": "2024",
        "요약": "Panel 채팅 기능을 활용한 다양한 예제 모음. LangChain, OpenAI, Mistral, Llama, RAG 구현 포함",
        "관련 섹션": "Panel LLM 통합 예시",
    },
    {
        "번호": 16,
        "제목": "Build a Mixtral Chatbot with Panel",
        "저자/출처": "HoloViz Blog",
        "URL": "https://blog.holoviz.org/posts/mixtral/",
        "발행일": "2024",
        "요약": "Panel을 사용하여 Mixtral LLM과 연동하는 챗봇을 만드는 방법. ChatInterface와 스트리밍 처리 설명",
        "관련 섹션": "Panel LLM 통합",
    },
    {
        "번호": 17,
        "제목": "GitHub - mesop-dev/mesop: Rapidly build AI apps in Python",
        "저자/출처": "Mesop GitHub (Google)",
        "URL": "https://github.com/mesop-dev/mesop",
        "발행일": "2024-2026",
        "요약": "Google Mesop 공식 GitHub 저장소. Python 기반 UI 프레임워크. 6.5k 스타, Apache 2.0 라이선스",
        "관련 섹션": "Mesop 개요",
    },
    {
        "번호": 18,
        "제목": "Google Mesop: An Open-source and Python-based UI Framework",
        "저자/출처": "Generative AI Pub",
        "URL": "https://www.generativeaipub.com/p/google-mesop-an-open-source-and-python",
        "발행일": "2024",
        "요약": "Google Mesop 프레임워크 소개. LLM 앱 개발에 최적화된 Python UI 프레임워크. 선언형 UI, Angular 기반, 핫 리로드",
        "관련 섹션": "Mesop 특징",
    },
    {
        "번호": 19,
        "제목": "Building a Simple GenAI Chatbot with Google Mesop",
        "저자/출처": "Medium - Ingrid Stevens",
        "URL": "https://medium.com/@ingridwickstevens/building-a-simple-genai-chatbot-with-google-mesop-e956e944d1b7",
        "발행일": "2024",
        "요약": "Google Mesop으로 생성형 AI 챗봇을 만드는 단계별 튜토리얼. mel.chat 컴포넌트 활용",
        "관련 섹션": "Mesop 챗봇 구현",
    },
    {
        "번호": 20,
        "제목": "GitHub - voila-dashboards/voila",
        "저자/출처": "Voilà GitHub",
        "URL": "https://github.com/voila-dashboards/voila",
        "발행일": "2024",
        "요약": "Voilà 공식 GitHub 저장소. Jupyter 노트북을 독립적인 웹앱으로 변환하는 도구. BSD 라이선스",
        "관련 섹션": "Voilà 개요",
    },
    {
        "번호": 21,
        "제목": "Deploying Voilà - Official Documentation",
        "저자/출처": "Voilà 공식 문서",
        "URL": "https://voila.readthedocs.io/en/stable/deploy.html",
        "발행일": "2024",
        "요약": "Voilà 앱 배포 가이드. Binder, Heroku, Google App Engine, JupyterHub, Ploomber Cloud 등 배포 옵션 설명",
        "관련 섹션": "Voilà 배포",
    },
    {
        "번호": 22,
        "제목": "Streamlit vs Gradio in 2025: A Framework Comparison for AI Apps",
        "저자/출처": "Squadbase Blog",
        "URL": "https://www.squadbase.dev/en/blog/streamlit-vs-gradio-in-2025-a-framework-comparison-for-ai-apps",
        "발행일": "2025",
        "요약": "2025년 기준 Streamlit과 Gradio의 심층 비교. UI 커스터마이징, 추론 관리, 배포, 인증 기능 비교",
        "관련 섹션": "프레임워크 비교",
    },
    {
        "번호": 23,
        "제목": "The 3 Best Python Frameworks To Build UIs for AI Apps",
        "저자/출처": "GetStream Blog",
        "URL": "https://getstream.io/blog/ai-chat-ui-tools/",
        "발행일": "2025",
        "요약": "Gradio, Streamlit, Chainlit 세 프레임워크를 코드 예시와 함께 비교. 선택 가이드와 사용 사례 포함",
        "관련 섹션": "프레임워크 비교, 코드 예시",
    },
    {
        "번호": 24,
        "제목": "Streamlit vs Gradio vs Chainlit: Best UI Framework for LLMs in 2025",
        "저자/출처": "Markaicode",
        "URL": "https://markaicode.com/streamlit-vs-gradio-vs-chainlit-llm-ui-framework/",
        "발행일": "2025",
        "요약": "LLM UI 프레임워크 3종 비교. 각 프레임워크의 강점, 약점, 적합한 사용 사례 정리",
        "관련 섹션": "프레임워크 비교",
    },
    {
        "번호": 25,
        "제목": "Building a Multimodal Gradio Chatbot with Llama 3.2 Using the Ollama API",
        "저자/출처": "PyImageSearch",
        "URL": "https://pyimagesearch.com/2025/02/10/building-a-multimodal-gradio-chatbot-with-llama-3-2-using-the-ollama-api/",
        "발행일": "2025-02",
        "요약": "Gradio + Ollama + Llama 3.2 조합으로 멀티모달 챗봇을 구현하는 방법. 이미지 처리 포함",
        "관련 섹션": "Gradio 멀티모달",
    },
    {
        "번호": 26,
        "제목": "Gradio + LiteLLM Tutorial",
        "저자/출처": "LiteLLM 공식 문서",
        "URL": "https://docs.litellm.ai/docs/tutorials/gradio_integration",
        "발행일": "2024",
        "요약": "Gradio와 LiteLLM을 통합하여 100개 이상의 LLM Provider를 지원하는 챗봇 구현 방법",
        "관련 섹션": "Gradio LLM 연동",
    },
    {
        "번호": 27,
        "제목": "A Survey of Python Frameworks - Ploomber Blog",
        "저자/출처": "Ploomber",
        "URL": "https://ploomber.io/blog/survey-python-frameworks/",
        "발행일": "2024",
        "요약": "파이썬 웹앱 프레임워크에 대한 포괄적인 서베이. Streamlit, Gradio, Panel, Dash 등 비교",
        "관련 섹션": "전체 프레임워크 비교",
    },
    {
        "번호": 28,
        "제목": "Build an AI Chatbot to Run Code and Tweak plots - HoloViz Blog",
        "저자/출처": "HoloViz Blog",
        "URL": "https://blog.holoviz.org/posts/tweak-mpl-chat/",
        "발행일": "2024",
        "요약": "Panel을 사용하여 코드 실행과 그래프 수정이 가능한 AI 챗봇을 만드는 방법. 데이터 과학 활용 사례",
        "관련 섹션": "Panel 고급 활용",
    },
    {
        "번호": 29,
        "제목": "Create a Simple ChatBot with Mesop + Ollama less than 25 lines",
        "저자/출처": "DEV Community",
        "URL": "https://dev.to/0xkoji/create-a-simple-chatbot-with-mesop-ollama-less-than-25-lines-2c9l",
        "발행일": "2024",
        "요약": "Mesop + Ollama로 25줄 이하의 코드로 로컬 LLM 챗봇을 만드는 빠른 가이드",
        "관련 섹션": "Mesop 코드 예시",
    },
    {
        "번호": 30,
        "제목": "And Voilà! From Jupyter notebooks to standalone web apps",
        "저자/출처": "Jupyter Blog - QuantStack",
        "URL": "https://blog.jupyter.org/and-voil%C3%A0-f6a2c08a4a93",
        "발행일": "2019 (원본), 지속 업데이트",
        "요약": "Voilà 프레임워크를 소개하는 공식 Jupyter Blog 포스트. 설계 철학과 기술적 구현 설명",
        "관련 섹션": "Voilà 배경",
    },
]


def create_xlsx():
    """참고문헌 XLSX 파일 생성"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # ─── 헤더 스타일 정의 ───────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="맑은 고딕")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 교대 행 색상
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # 테두리 스타일
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ─── 헤더 행 작성 ───────────────────────────────────────────
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # ─── 데이터 행 작성 ─────────────────────────────────────────
    for row_idx, ref in enumerate(REFERENCES, start=2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        row_data = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border

            if col_idx == 1:  # 번호: 가운데 정렬
                cell.alignment = center_align
                cell.font = Font(bold=True, size=10, name="맑은 고딕")
            elif col_idx == 4:  # URL: 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = Font(color="0563C1", underline="single", size=9, name="맑은 고딕")
                cell.alignment = left_align
            else:
                cell.alignment = left_align
                cell.font = Font(size=10, name="맑은 고딕")

    # ─── 열 너비 설정 ───────────────────────────────────────────
    column_widths = {
        "A": 6,    # 번호
        "B": 45,   # 제목
        "C": 22,   # 저자/출처
        "D": 55,   # URL
        "E": 12,   # 발행일
        "F": 60,   # 요약
        "G": 25,   # 관련 섹션
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ─── 행 높이 설정 ───────────────────────────────────────────
    ws.row_dimensions[1].height = 30  # 헤더 행
    for row_idx in range(2, len(REFERENCES) + 2):
        ws.row_dimensions[row_idx].height = 55

    # ─── 고정 창 설정 (헤더 행 고정) ──────────────────────────
    ws.freeze_panes = "A2"

    # ─── 자동 필터 ───────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # ─── 두 번째 시트: 요약 통계 ────────────────────────────────
    ws_summary = wb.create_sheet(title="관련 섹션별 분류")

    # 섹션별 분류
    section_map = {}
    for ref in REFERENCES:
        sections = [s.strip() for s in ref["관련 섹션"].split(",")]
        for section in sections:
            if section not in section_map:
                section_map[section] = []
            section_map[section].append(ref["번호"])

    # 요약 시트 헤더
    ws_summary["A1"] = "관련 섹션"
    ws_summary["B1"] = "참고문헌 번호"
    ws_summary["C1"] = "수량"

    ws_summary["A1"].fill = header_fill
    ws_summary["A1"].font = header_font
    ws_summary["B1"].fill = header_fill
    ws_summary["B1"].font = header_font
    ws_summary["C1"].fill = header_fill
    ws_summary["C1"].font = header_font

    for row_idx, (section, nums) in enumerate(sorted(section_map.items()), start=2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        ws_summary.cell(row=row_idx, column=1, value=section).fill = row_fill
        ws_summary.cell(row=row_idx, column=2, value=", ".join(map(str, nums))).fill = row_fill
        ws_summary.cell(row=row_idx, column=3, value=len(nums)).fill = row_fill

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 8

    # ─── 파일 저장 ───────────────────────────────────────────────
    output_path = "python-chatbot-tools-references.xlsx"
    wb.save(output_path)
    print(f"참고문헌 파일 생성 완료: {output_path}")
    print(f"총 {len(REFERENCES)}개의 참고문헌")


if __name__ == "__main__":
    create_xlsx()
