"""
Chainlit 리서치 - 참고문헌 XLSX 생성 스크립트
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# 참고문헌 데이터
references = [
    {
        "번호": 1,
        "제목": "Chainlit 공식 문서 - 개요",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/get-started/overview",
        "발행일": "2025-2026",
        "요약": "Chainlit의 공식 시작 가이드. 프레임워크 개요, 주요 기능(빠른 개발, 인증, 데이터 지속성, 다단계 추론 시각화, 멀티플랫폼), 통합 지원 목록 포함.",
        "관련 섹션": "1. 개요, 3. 아키텍처"
    },
    {
        "번호": 2,
        "제목": "Chainlit GitHub 리포지토리",
        "저자/출처": "Chainlit/chainlit (GitHub)",
        "URL": "https://github.com/Chainlit/chainlit",
        "발행일": "2026-03-05",
        "요약": "Chainlit 오픈소스 리포지토리. 최신 버전 2.10.0, 스타 11.7k, 포크 1.7k, Apache 2.0 라이선스. 2025년 5월부터 커뮤니티 유지보수 전환.",
        "관련 섹션": "1. 개요, 9. 트렌드"
    },
    {
        "번호": 3,
        "제목": "Chainlit 공식 홈페이지",
        "저자/출처": "Chainlit",
        "URL": "https://chainlit.io/",
        "발행일": "2025-2026",
        "요약": "'Build production ready Conversational AI in minutes, not weeks' - Chainlit 브랜드 슬로건 및 제품 소개 페이지.",
        "관련 섹션": "1. 개요"
    },
    {
        "번호": 4,
        "제목": "chainlit PyPI 패키지",
        "저자/출처": "PyPI",
        "URL": "https://pypi.org/project/chainlit/",
        "발행일": "2026-03",
        "요약": "Chainlit Python 패키지 공식 배포 페이지. 버전 이력, 의존성, 설치 방법 포함. 최신 버전 2.10.0.",
        "관련 섹션": "2. 설치, 9. 트렌드"
    },
    {
        "번호": 5,
        "제목": "Chainlit 사용자 세션 문서",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/concepts/user-session",
        "발행일": "2025-2026",
        "요약": "cl.user_session API 공식 문서. 멀티유저 데이터 격리, set/get 메서드 사용법, 예약된 키 목록, 전역 변수 사용 금지 이유 설명.",
        "관련 섹션": "3. 아키텍처, 4. 핵심 기능"
    },
    {
        "번호": 6,
        "제목": "Chainlit 멀티모달 공식 문서",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/advanced-features/multi-modal",
        "발행일": "2025-2026",
        "요약": "Chainlit의 멀티모달 지원 기능 문서. 파일 업로드(드래그&드롭), 이미지 필터링, 음성 스트림(@cl.on_audio_chunk), OpenAI Realtime API 통합 방법.",
        "관련 섹션": "4. 핵심 기능 - 멀티모달"
    },
    {
        "번호": 7,
        "제목": "Chainlit LangChain/LangGraph 통합 문서",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/integrations/langchain",
        "발행일": "2025-2026",
        "요약": "LangChain LCEL과 LangGraph를 Chainlit에 통합하는 방법. cl.LangchainCallbackHandler 사용법, LCEL 체인 구성, astream 스트리밍 코드 예제.",
        "관련 섹션": "6. LangChain/LangGraph 통합"
    },
    {
        "번호": 8,
        "제목": "Chainlit 배포 공식 문서",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/deploy/overview",
        "발행일": "2025-2026",
        "요약": "Chainlit 프로덕션 배포 가이드. 지원 플랫폼(Web App, Copilot, Teams/Slack/Discord), Docker 배포, WebSocket sticky session 설정, CORS 설정.",
        "관련 섹션": "8. 프로덕션 배포"
    },
    {
        "번호": 9,
        "제목": "DeepWiki - Chainlit 세션 관리 아키텍처",
        "저자/출처": "DeepWiki",
        "URL": "https://deepwiki.com/Chainlit/chainlit/3.4-session-management",
        "발행일": "2025",
        "요약": "Chainlit 세션 관리 내부 아키텍처 심층 분석. ws_sessions_sid/ws_sessions_id 딕셔너리, 세션 복구 메커니즘, 지연 정리(deferred cleanup) 방식 설명.",
        "관련 섹션": "3. 아키텍처, 4. 세션 관리"
    },
    {
        "번호": 10,
        "제목": "DeepWiki - Chainlit 인증 시스템",
        "저자/출처": "DeepWiki",
        "URL": "https://deepwiki.com/Chainlit/chainlit/8-authentication",
        "발행일": "2025",
        "요약": "Chainlit 인증 시스템 심층 분석. 패스워드 인증, JWT 토큰, OAuth 제공자 통합, 사용자 세션과의 연동 방식.",
        "관련 섹션": "4. 인증 기능"
    },
    {
        "번호": 11,
        "제목": "The 3 Best Python Frameworks To Build UIs for AI Apps",
        "저자/출처": "getstream.io",
        "URL": "https://getstream.io/blog/ai-chat-ui-tools/",
        "발행일": "2025",
        "요약": "Chainlit, Streamlit, Gradio 비교 분석. 각 프레임워크의 특징, 장단점, 모니터링/커스터마이징/배포 비교표, 용도별 추천 가이드.",
        "관련 섹션": "5. 프레임워크 비교"
    },
    {
        "번호": 12,
        "제목": "Streamlit vs Gradio vs Chainlit: Best UI Framework for LLMs in 2025",
        "저자/출처": "markaicode.com",
        "URL": "https://markaicode.com/streamlit-vs-gradio-vs-chainlit-llm-ui-framework/",
        "발행일": "2025",
        "요약": "LLM UI 프레임워크 3종 비교. 학습 곡선, 채팅 UI 품질, 스트리밍, 파일 업로드, 인증, 모니터링, 멀티플랫폼 지원 등 종합 비교.",
        "관련 섹션": "5. 프레임워크 비교"
    },
    {
        "번호": 13,
        "제목": "Building a Simple Chatbot with LangGraph and Chainlit",
        "저자/출처": "James B Mour (DEV Community)",
        "URL": "https://dev.to/jamesbmour/building-a-simple-chatbot-with-langgraph-and-chainlit-a-step-by-step-tutorial-4k6h",
        "발행일": "2025",
        "요약": "LangGraph + Chainlit 통합 단계별 튜토리얼. StateGraph, MessagesState, MemorySaver 사용법, 스트리밍 응답 구현, 약 50줄로 완전한 챗봇 구현.",
        "관련 섹션": "6. LangGraph 통합"
    },
    {
        "번호": 14,
        "제목": "chainlit_langgraph - GitHub 리포지토리",
        "저자/출처": "brucechou1983 (GitHub)",
        "URL": "https://github.com/brucechou1983/chainlit_langgraph",
        "발행일": "2025",
        "요약": "Chainlit + LangGraph 통합 오픈소스 프로젝트. 다중 LLM 지원(Ollama, Claude, GPT 등), 워크플로우 패턴, BaseWorkflow 추상 클래스 구조.",
        "관련 섹션": "6. LangGraph 통합"
    },
    {
        "번호": 15,
        "제목": "RAG in Production with LangChain and Chainlit",
        "저자/출처": "Duy Nguyen (Medium)",
        "URL": "https://medium.com/@justinduy/rag-in-production-with-langchain-and-chainlit-86c2dea0ca40",
        "발행일": "2025",
        "요약": "프로덕션 RAG 구현 가이드. LangChain + Chainlit 조합으로 문서 검색 챗봇 구현, Ray를 이용한 확장성, NewRelic 모니터링 통합.",
        "관련 섹션": "7. 활용 사례, 8. 프로덕션 배포"
    },
    {
        "번호": 16,
        "제목": "It's 2025 - Start Using Chainlit for Your LLM Apps",
        "저자/출처": "Tituslhy (Medium / MITB For All)",
        "URL": "https://medium.com/mitb-for-all/its-2025-start-using-chainlit-for-your-llm-apps-558db1a46315",
        "발행일": "2025",
        "요약": "2025년 기준 Chainlit 입문 가이드. 설치부터 고급 기능까지 단계별 소개, 실습 예제 포함.",
        "관련 섹션": "2. 설치, 4. 핵심 기능"
    },
    {
        "번호": 17,
        "제목": "Chainlit AI Framework Flaws Enable Data Theft via File Read and SSRF Bugs",
        "저자/출처": "The Hacker News",
        "URL": "https://thehackernews.com/2026/01/chainlit-ai-framework-flaws-enable-data.html",
        "발행일": "2026-01",
        "요약": "2026년 1월 발견된 Chainlit 보안 취약점 보고. 파일 읽기 및 SSRF 버그를 통한 데이터 탈취 가능성 경고, 최신 버전 업데이트 권장.",
        "관련 섹션": "9. 트렌드 및 보안"
    },
    {
        "번호": 18,
        "제목": "Building a Weather Agent with LangGraph, Chainlit & MCP",
        "저자/출처": "Dominic Schneider (Medium)",
        "URL": "https://medium.com/@dominicschneider_7223/%EF%B8%8F-building-a-weather-agent-with-langgraph-chainlit-mcp-your-first-modular-ai-tool-6208bbb3d693",
        "발행일": "2025",
        "요약": "LangGraph + Chainlit + MCP(Model Context Protocol) 조합으로 날씨 에이전트 구현 튜토리얼. FastMCP와의 통합 방법 포함.",
        "관련 섹션": "6. LangGraph 통합, 9. 트렌드"
    },
    {
        "번호": 19,
        "제목": "Chat with your Data: Building a File-Aware AI Agent with AWS Bedrock and Chainlit",
        "저자/출처": "gonzalo123.com",
        "URL": "https://gonzalo123.com/2025/12/09/chat-with-your-data-building-a-file-aware-ai-agent-with-aws-bedrock-and-chainlit/",
        "발행일": "2025-12",
        "요약": "AWS Bedrock + Chainlit 통합으로 파일 인식 AI 에이전트 구현. 기업 AWS 환경에서의 Chainlit 활용 사례.",
        "관련 섹션": "7. 활용 사례, 9. 트렌드"
    },
    {
        "번호": 20,
        "제목": "Building a News and Stock Assistant with LangGraph and Chainlit",
        "저자/출처": "James B Mour (DEV Community)",
        "URL": "https://dev.to/jamesbmour/building-a-news-and-stock-assistant-with-langgraph-and-chainlit-1bkk",
        "발행일": "2025",
        "요약": "LangGraph + Chainlit으로 뉴스 및 주식 정보 조회 어시스턴트 구현. 다중 도구 호출, 조건부 라우팅, 실시간 데이터 통합.",
        "관련 섹션": "6. LangGraph 통합, 7. 활용 사례"
    },
    {
        "번호": 21,
        "제목": "Chainlit: A Guide With Practical Examples",
        "저자/출처": "DataCamp",
        "URL": "https://www.datacamp.com/tutorial/chainlit",
        "발행일": "2025",
        "요약": "DataCamp의 Chainlit 실습 가이드. 기본 설치부터 파일 업로드, 세션 관리, 인증, 데이터 레이어까지 단계별 코드 예제 포함.",
        "관련 섹션": "2. 설치, 4. 핵심 기능"
    },
    {
        "번호": 22,
        "제목": "Chainlit: The Developer's Guide to Building Conversational AI",
        "저자/출처": "Skywork AI",
        "URL": "https://skywork.ai/skypage/en/Chainlit:-The-Developer's-Guide-to-Building-Conversational-AI/1976187694951886848",
        "발행일": "2025",
        "요약": "Chainlit 개발자 종합 가이드. 하이브리드 아키텍처(Python 백엔드 + React 프론트엔드), 통합 지원 프레임워크 목록, 실제 구현 패턴.",
        "관련 섹션": "3. 아키텍처"
    },
    {
        "번호": 23,
        "제목": "DeepWiki - Chainlit 백엔드 아키텍처",
        "저자/출처": "DeepWiki",
        "URL": "https://deepwiki.com/Chainlit/chainlit/2.1-backend-architecture",
        "발행일": "2025",
        "요약": "Chainlit 백엔드 아키텍처 심층 분석. WebSocket 기반 통신, 이벤트 드리븐 구조, 비동기 처리 방식.",
        "관련 섹션": "3. 아키텍처"
    },
    {
        "번호": 24,
        "제목": "Migrate to Chainlit v2.0.0",
        "저자/출처": "Chainlit",
        "URL": "https://docs.chainlit.io/guides/migration/2.0.0",
        "발행일": "2025",
        "요약": "Chainlit v1.x에서 v2.0.0으로 마이그레이션 공식 가이드. @data_layer 데코레이터, Shadcn/Tailwind UI 재작성, API 변경사항.",
        "관련 섹션": "9. 트렌드"
    },
    {
        "번호": 25,
        "제목": "Mesop, Streamlit, Chainlit, and Gradio: A Comprehensive Comparison",
        "저자/출처": "AI Hive",
        "URL": "https://www.ai-hive.net/post/mesop-streamlit-chainlit-and-gradio-a-comprehensive-comparison-of-ai-application-frameworks",
        "발행일": "2025",
        "요약": "4개 AI 앱 프레임워크 종합 비교. Mesop(Google), Streamlit, Chainlit, Gradio의 기능, 성능, 생태계 비교 분석.",
        "관련 섹션": "5. 프레임워크 비교"
    },
    {
        "번호": 26,
        "제목": "Guide to Deploying Chainlit with RAG on Upsun",
        "저자/출처": "DEV Community / Upsun",
        "URL": "https://dev.to/upsun/guide-deploying-chainlit-with-rag-on-upsun-2f9c",
        "발행일": "2025",
        "요약": "Upsun 클라우드 플랫폼에 Chainlit + RAG 배포 가이드. 컨테이너 설정, 환경 변수 관리, 프로덕션 최적화.",
        "관련 섹션": "8. 프로덕션 배포"
    },
    {
        "번호": 27,
        "제목": "LangGraph Agent Orchestration Framework",
        "저자/출처": "LangChain",
        "URL": "https://www.langchain.com/langgraph",
        "발행일": "2025-2026",
        "요약": "LangGraph 공식 소개 페이지. 에이전트 오케스트레이션, 상태 그래프 관리, 체크포인팅, 멀티 에이전트 패턴 설명.",
        "관련 섹션": "6. LangGraph 통합"
    },
    {
        "번호": 28,
        "제목": "Building Multi-Agent Systems with LangGraph: A Step-by-Step Guide",
        "저자/출처": "Sushmita Nandi (Medium)",
        "URL": "https://medium.com/@sushmita2310/building-multi-agent-systems-with-langgraph-a-step-by-step-guide-d14088e90f72",
        "발행일": "2025",
        "요약": "LangGraph 멀티 에이전트 시스템 구현 가이드. create_react_agent 패턴, 슈퍼바이저 에이전트, 상태 공유, 메시지 라우팅.",
        "관련 섹션": "6. LangGraph 통합"
    },
    {
        "번호": 29,
        "제목": "Deploying chainlit application using NGINX",
        "저자/출처": "Palash Chandra (Medium)",
        "URL": "https://medium.com/@palash-fin/deploying-chainlit-application-using-nginx-c95069360983",
        "발행일": "2025",
        "요약": "NGINX 리버스 프록시를 이용한 Chainlit 배포 가이드. WebSocket 설정, SSL 인증서, 로드밸런싱 구성.",
        "관련 섹션": "8. 프로덕션 배포"
    },
    {
        "번호": 30,
        "제목": "My First Time Building a Chainlit Data Layer",
        "저자/출처": "Andres Felipe Tellez Yepes (Medium)",
        "URL": "https://medium.com/@andres.tellez/my-first-time-building-a-chainlit-data-layer-dd6de99b6b80",
        "발행일": "2025",
        "요약": "Chainlit 커스텀 데이터 레이어 구현 경험담. BaseDataLayer 상속, DB 연동, 대화 기록 영속성 구현 방법.",
        "관련 섹션": "4. 데이터 영속성"
    },
]


def create_references_xlsx():
    """참고문헌 XLSX 파일을 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 컬러 설정
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ROW_FILL_1 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ROW_FILL_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    HEADER_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Malgun Gothic", size=10)
    URL_FONT = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    BOLD_FONT = Font(name="Malgun Gothic", bold=True, size=10)

    THIN_BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 정의
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 40, 20, 45, 10, 60, 25]

    # 헤더 행 작성
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 데이터 행 작성
    for row_idx, ref in enumerate(references, 2):
        is_odd = (row_idx % 2 == 0)
        row_fill = ROW_FILL_1 if is_odd else ROW_FILL_2

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER

            # 번호 열
            if col_idx == 1:
                cell.font = BOLD_FONT
                cell.alignment = CENTER
            # URL 열
            elif col_idx == 4:
                cell.font = URL_FONT
                cell.alignment = LEFT
                cell.hyperlink = value
            # 요약 열
            elif col_idx == 6:
                cell.font = BODY_FONT
                cell.alignment = LEFT
            else:
                cell.font = BODY_FONT
                cell.alignment = LEFT

        # 행 높이 설정
        ws.row_dimensions[row_idx].height = 55

    # 헤더 행 높이
    ws.row_dimensions[1].height = 25

    # 틀 고정 (헤더 행)
    ws.freeze_panes = "A2"

    # 자동 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 저장
    output_path = "chainlit-references.xlsx"
    wb.save(output_path)
    print(f"[완료] 참고문헌 파일 생성: {output_path}")
    print(f"[정보] 총 {len(references)}개 참고문헌 포함")
    return output_path


if __name__ == "__main__":
    create_references_xlsx()
