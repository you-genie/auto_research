"""
참고문헌 XLSX 파일 생성 스크립트
Svelte 가이드 101 — 온보딩 클래스 (Svelte 5 Runes 기준, 2025~2026)
"""

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

REFERENCES = [
    {
        "번호": 1,
        "제목": "Introducing runes",
        "저자/출처": "Svelte Blog",
        "URL": "https://svelte.dev/blog/runes",
        "발행일": "2024-01-01",
        "요약 (한국어)": "Svelte 5 Runes 도입 공식 발표. $state/$derived/$effect가 평가 시점에 의존성을 판단하는 명시적 반응성으로, Svelte 4의 컴파일러 기반 암묵 반응성을 대체하는 배경과 설계 철학.",
        "관련 섹션": "반응성의 핵심 - Runes",
    },
    {
        "번호": 2,
        "제목": "$state — Svelte Docs",
        "저자/출처": "Svelte (공식 문서)",
        "URL": "https://svelte.dev/docs/svelte/$state",
        "발행일": "2025-01-01",
        "요약 (한국어)": "$state rune 공식 문서. 반응형 상태 선언, 객체·배열의 깊은(deep) 반응성, $state.raw 등 변형. let 기반 암묵 반응성을 대체.",
        "관련 섹션": "Runes - $state",
    },
    {
        "번호": 3,
        "제목": "$derived — Svelte Docs",
        "저자/출처": "Svelte (공식 문서)",
        "URL": "https://svelte.dev/docs/svelte/$derived",
        "발행일": "2025-01-01",
        "요약 (한국어)": "$derived rune 공식 문서. 의존 상태 변경 시 자동 재계산되는 메모이제이션 파생값. 복잡한 로직을 위한 $derived.by 사용법.",
        "관련 섹션": "Runes - $derived",
    },
    {
        "번호": 4,
        "제목": "$effect — Svelte Docs",
        "저자/출처": "Svelte (공식 문서)",
        "URL": "https://svelte.dev/docs/svelte/$effect",
        "발행일": "2025-01-01",
        "요약 (한국어)": "$effect rune 공식 문서. 마운트·의존성 변경 시 실행되는 부수 효과. DOM 조작·서드파티 연동 등 외부 동기화에 한정 사용하라는 가이드.",
        "관련 섹션": "Runes - $effect",
    },
    {
        "번호": 5,
        "제목": "Understanding Svelte 5 Runes: $derived vs $effect",
        "저자/출처": "DEV Community (Mike)",
        "URL": "https://dev.to/mikehtmlallthethings/understanding-svelte-5-runes-derived-vs-effect-1hh",
        "발행일": "2025-01-01",
        "요약 (한국어)": "$derived와 $effect의 차이와 선택 기준 해설. 계산되는 값은 derived, 외부 동기화만 effect라는 핵심 규칙과 effect 남용 시 발생하는 문제를 설명.",
        "관련 섹션": "Runes - derived vs effect",
    },
    {
        "번호": 6,
        "제목": "Svelte 5 runes — the complete guide",
        "저자/출처": "Full Stack SvelteKit",
        "URL": "https://fullstacksveltekit.com/blog/svelte-5-runes",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Svelte 5 Runes 종합 가이드. $state/$derived/$effect/$props와 $bindable·$inspect 등 전체 rune을 예제 중심으로 정리.",
        "관련 섹션": "Runes 전반",
    },
    {
        "번호": 7,
        "제목": "Svelte 5 Runes in 2026: How They Work",
        "저자/출처": "PkgPulse Blog",
        "URL": "https://www.pkgpulse.com/blog/svelte-5-runes-complete-guide-2026",
        "발행일": "2026-01-01",
        "요약 (한국어)": "2026 기준 Runes 동작 원리 가이드. 반응성 그래프, 의존성 추적, 메모이제이션 메커니즘을 최신 관점에서 설명.",
        "관련 섹션": "Runes 심화",
    },
    {
        "번호": 8,
        "제목": "Routing — SvelteKit Docs",
        "저자/출처": "SvelteKit (공식 문서)",
        "URL": "https://svelte.dev/docs/kit/routing",
        "발행일": "2025-01-01",
        "요약 (한국어)": "SvelteKit 파일 기반 라우팅 공식 문서. +page/+layout/+error 등 + 접두 파일 규칙, [slug] 동적 파라미터, load 함수와 서버 전용 +page.server.js 구분.",
        "관련 섹션": "SvelteKit - 라우팅",
    },
    {
        "번호": 9,
        "제목": "Svelte Tutorial (공식 인터랙티브)",
        "저자/출처": "Svelte",
        "URL": "https://svelte.dev/tutorial",
        "발행일": "2025-01-01",
        "요약 (한국어)": "브라우저에서 바로 실습하는 공식 인터랙티브 튜토리얼. Svelte 기초부터 SvelteKit까지 단계별 실습 — 온보딩 클래스 최고의 시작점.",
        "관련 섹션": "학습 리소스",
    },
    {
        "번호": 10,
        "제목": "Routing / Pages — Svelte Tutorial",
        "저자/출처": "Svelte (공식 튜토리얼)",
        "URL": "https://svelte.dev/tutorial/kit/pages",
        "발행일": "2025-01-01",
        "요약 (한국어)": "SvelteKit 페이지·라우팅 실습 튜토리얼. src/routes에 +page.svelte를 두면 페이지가 생성되는 파일 기반 라우팅을 손으로 익힌다.",
        "관련 섹션": "SvelteKit - 페이지",
    },
    {
        "번호": 11,
        "제목": "SvelteKit Routing Tutorial: Layouts, Nested Routes & Multi-Page Apps",
        "저자/출처": "DEV Community",
        "URL": "https://dev.to/a1guy/sveltekit-routing-tutorial-layouts-nested-routes-multi-page-apps-4bfm",
        "발행일": "2025-01-01",
        "요약 (한국어)": "레이아웃·중첩 라우트·멀티페이지 앱 구성 튜토리얼. +layout.svelte로 공통 구조를 감싸고 중첩 라우트를 만드는 실무 패턴.",
        "관련 섹션": "SvelteKit - 레이아웃",
    },
    {
        "번호": 12,
        "제목": "SvelteKit 2.0 Routing and Layouts: Best Practices",
        "저자/출처": "Java Code Geeks",
        "URL": "https://www.javacodegeeks.com/2025/05/sveltekit-2-0-routing-and-layouts-best-practices-for-building-maintainable-apps.html",
        "발행일": "2025-05-01",
        "요약 (한국어)": "SvelteKit 2.0 라우팅·레이아웃 베스트 프랙티스. 유지보수 가능한 앱 구조를 위한 폴더 설계와 load/actions 조직화 원칙.",
        "관련 섹션": "SvelteKit - 베스트 프랙티스",
    },
    {
        "번호": 13,
        "제목": "Svelte 5 Cheat Sheet — Complete SvelteKit Reference",
        "저자/출처": "devtooleasy",
        "URL": "https://devtooleasy.com/cheat-sheet/svelte",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Svelte 5 + SvelteKit 치트시트. Runes, 템플릿 문법, 이벤트, 바인딩, 라우팅을 한눈에 정리 — 수강생 배포용 참고 자료.",
        "관련 섹션": "치트시트",
    },
    {
        "번호": 14,
        "제목": "6 Svelte 5 Runes That Make Refs & Stores Obvious",
        "저자/출처": "Medium (Neurobyte)",
        "URL": "https://medium.com/@kaushalsinh73/6-svelte-5-runes-that-make-refs-stores-obvious-18be2b34bdb9",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Runes로 ref·store 개념이 어떻게 단순해지는지 설명. $bindable, $props 등으로 컴포넌트 통신과 참조를 다루는 실전 예제.",
        "관련 섹션": "Props와 컴포넌트 통신",
    },
    {
        "번호": 15,
        "제목": "Svelte 5 Refresher with Runes",
        "저자/출처": "luminary.blog",
        "URL": "https://luminary.blog/techs/05-svelte5-refresher/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Svelte 4 경험자를 위한 5 리프레셔. 이벤트(onclick), Snippets, 콜백 prop 등 4→5 변경점을 빠르게 훑는 마이그레이션 관점 정리.",
        "관련 섹션": "Svelte 4 → 5 마이그레이션",
    },
    {
        "번호": 16,
        "제목": "Svelte 5 Docs — Overview",
        "저자/출처": "Svelte (공식 문서)",
        "URL": "https://svelte.dev/docs/svelte/overview",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Svelte 5 공식 문서 개요. 컴포넌트 구조, Runes, 템플릿 문법, 스타일, 트랜지션 등 전체 레퍼런스의 진입점.",
        "관련 섹션": "컴포넌트 구조 / 전반",
    },
    {
        "번호": 17,
        "제목": "SvelteKit Docs — Introduction",
        "저자/출처": "SvelteKit (공식 문서)",
        "URL": "https://svelte.dev/docs/kit/introduction",
        "발행일": "2025-01-01",
        "요약 (한국어)": "SvelteKit 소개 문서. 메타프레임워크 개념, SSR/SSG/CSR/ISR 렌더링 모드, 프로젝트 생성(npx sv create)과 기본 구조.",
        "관련 섹션": "SvelteKit - 개요",
    },
    {
        "번호": 18,
        "제목": "Runes in Svelte 5",
        "저자/출처": "CodeHints",
        "URL": "https://codehints.io/svelte/runes/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Runes 입문 가이드. $state/$derived/$effect/$props의 역할과 사용 예를 간결하게 정리한 초보자용 레퍼런스.",
        "관련 섹션": "Runes 입문",
    },
]


def create_xlsx(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 스타일 정의 (Svelte 오렌지 테마)
    header_fill = PatternFill(start_color="FF3E00", end_color="FF3E00", fill_type="solid")
    even_fill = PatternFill(start_color="FDE3D8", end_color="FDE3D8", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Malgun Gothic", size=10)
    link_font = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 행
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 45, 25, 50, 12, 60, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, ref in enumerate(REFERENCES, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약 (한국어)"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            ws.row_dimensions[row_idx].height = 60

            if col_idx == 1:  # 번호
                cell.alignment = center_align
                cell.font = Font(name="Malgun Gothic", size=10, bold=True)
            elif col_idx == 4:  # URL — 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = link_font
                cell.alignment = left_align
            elif col_idx == 5:  # 발행일
                cell.alignment = center_align
                cell.font = body_font
            else:
                cell.alignment = left_align
                cell.font = body_font

    # 틀 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "FF3E00"

    wb.save(output_path)
    print(f"XLSX 파일 생성 완료: {output_path}")
    print(f"총 참고문헌 수: {len(REFERENCES)}개")


if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "svelte-guide-101-references.xlsx")
    create_xlsx(output_path)
