"""
LLM-as-a-Verifier 리서치 — 참고문헌 XLSX 생성 스크립트
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


references = [
    # Foundations (원조 계보)
    {
        "번호": 1,
        "제목": "Training Verifiers to Solve Math Word Problems",
        "저자/출처": "Cobbe et al. (OpenAI)",
        "URL": "https://arxiv.org/abs/2110.14168",
        "발행일": "2021-10",
        "요약": "LLM-as-a-verifier의 사실상 원점. GPT-3-6B 세대에서 정답 여부를 판정하는 verifier를 별도 학습하고 test-time에 100 후보를 샘플링해 verifier로 re-rank(Best-of-N). GSM8K 벤치마크(8.5K)를 함께 공개.",
        "관련 섹션": "3. 원조 계보"
    },
    {
        "번호": 2,
        "제목": "Solving math word problems with process- and outcome-based feedback",
        "저자/출처": "Uesato et al. (DeepMind)",
        "URL": "https://arxiv.org/abs/2211.14275",
        "발행일": "2022-11",
        "요약": "PRM/ORM 용어의 명명 기원. 각 reasoning step에 human annotation을 붙여 process-based feedback vs outcome-only feedback을 head-to-head 비교. 최종 답 정확도는 비슷하지만 trace error는 process가 확연히 낮춤.",
        "관련 섹션": "3. 원조 계보"
    },
    {
        "번호": 3,
        "제목": "Let's Verify Step by Step",
        "저자/출처": "Lightman et al. (OpenAI)",
        "URL": "https://arxiv.org/abs/2305.20050",
        "발행일": "2023-05",
        "요약": "Uesato의 가설을 대규모 스케일업. PRM800K(MATH 12K 문제 · 800K step 라벨) 공개. Best-of-N에서 PRM이 ORM 대비 큰 차이, GPT-4 + PRM으로 MATH 대표 서브셋 78%. o1/o3 process supervision의 뿌리.",
        "관련 섹션": "3. 원조 계보"
    },
    {
        "번호": 4,
        "제목": "PRM800K dataset",
        "저자/출처": "OpenAI (GitHub)",
        "URL": "https://github.com/openai/prm800k",
        "발행일": "2023",
        "요약": "Lightman et al. 2023이 공개한 스텝 단위 human-labeled dataset. MATH 12K 문제에 대해 약 800K step-level correctness annotation.",
        "관련 섹션": "3. 원조 계보, 9. 벤치마크"
    },
    {
        "번호": 5,
        "제목": "Math-Shepherd: Verify and Reinforce LLMs Step-by-step without Human Annotations",
        "저자/출처": "Wang et al. (PKU / DeepSeek)",
        "URL": "https://arxiv.org/abs/2312.08935",
        "발행일": "2023-12",
        "요약": "Monte Carlo rollout으로 각 스텝의 process reward를 자동 라벨링 → PRM 학습 + step-level PPO. Mistral-7B GSM8K 77.9→84.1%, MATH 28.6→33.0%. 오픈소스 PRM 파이프라인의 표준 레시피.",
        "관련 섹션": "3. 원조 계보, 6. RLVR"
    },
    # Judge 대조군
    {
        "번호": 6,
        "제목": "Judging LLM-as-a-Judge with MT-Bench and Chatbot Arena",
        "저자/출처": "Zheng et al. (LMSYS)",
        "URL": "https://arxiv.org/abs/2306.05685",
        "발행일": "2023-06",
        "요약": "LLM-as-a-Judge 계보의 정착점. GPT-4 심판과 인간의 agreement가 ~85%로 human-human 상회. MT-Bench(80 multi-turn) + Chatbot Arena(pairwise Elo)라는 표준을 남김. Verifier 계보와의 대조로 인용.",
        "관련 섹션": "2. Judge vs Verifier"
    },
    # Generative verifier 세대
    {
        "번호": 7,
        "제목": "Generative Verifiers: Reward Modeling as Next-Token Prediction (GenRM)",
        "저자/출처": "Zhang et al. (Google DeepMind)",
        "URL": "https://arxiv.org/abs/2408.15240",
        "발행일": "2024-08",
        "요약": "Discriminative head 대신 next-token P('Yes'|context)를 score로 사용. GenRM-CoT는 verification rationale + majority vote로 test-time scaling. 알고리즘 태스크 5→45.3%, GSM8K 73→93.4%, MATH easy-to-hard 28→44.6%, 6.4× data-efficient.",
        "관련 섹션": "4. Generative Verifier"
    },
    {
        "번호": 8,
        "제목": "Critique-out-Loud Reward Models (CLoud)",
        "저자/출처": "Ankner et al. (Databricks / MIT)",
        "URL": "https://arxiv.org/abs/2408.11791",
        "발행일": "2024-08",
        "요약": "자연어 critique 생성 후 scalar reward head를 붙이는 hybrid. RLHF preference RM 세팅에서 Pareto improvement. GenRM(수학·논리 verifier)과 대비되는 preference RM 도메인에서 explicit CoT chain의 이득 증거.",
        "관련 섹션": "4. Generative Verifier"
    },
    {
        "번호": 9,
        "제목": "Self-Taught Evaluators",
        "저자/출처": "Wang et al. (Meta FAIR)",
        "URL": "https://arxiv.org/abs/2408.02666",
        "발행일": "2024-08",
        "요약": "LLM judge를 human preference data 없이 학습하는 iterative bootstrapping. Synthetic contrasting output 생성 → LLM-judge가 explanation + score → 재훈련. RewardBench에서 GPT-4 judge에 근접. Verifier self-play 가능성 실증.",
        "관련 섹션": "4. Generative Verifier"
    },
    {
        "번호": 10,
        "제목": "Process Reward Models That Think (ThinkPRM)",
        "저자/출처": "Khalifa et al. (UMich / LG AI / Google)",
        "URL": "https://arxiv.org/abs/2504.16828",
        "발행일": "2025-04",
        "요약": "GenRM을 step-level로 확장. 각 스텝마다 long CoT verification rationale 후 step-wise Yes/No. PRM800K 1% 라벨만으로 discriminative baseline(full-supervision) outperform. OOD(GPQA, LiveCodeBench)에서 +8%, +4.5%.",
        "관련 섹션": "4. Generative Verifier"
    },
    # RLVR
    {
        "번호": 11,
        "제목": "Tulu 3: Pushing Frontiers in Open Language Model Post-Training",
        "저자/출처": "Lambert et al. (Ai2)",
        "URL": "https://arxiv.org/abs/2411.15124",
        "발행일": "2024-11",
        "요약": "RLVR(Reinforcement Learning with Verifiable Rewards) 프레임의 정식 명명. PPO/GRPO objective 유지 + 학습된 RM을 결정론적 verify() 함수로 대체. 정답 매칭 · IFEval 정규식 · 유닛테스트 등을 리워드로. 8B/70B 오픈.",
        "관련 섹션": "6. RLVR"
    },
    {
        "번호": 12,
        "제목": "DeepSeek-R1: Incentivizing Reasoning Capability in LLMs via Reinforcement Learning",
        "저자/출처": "DeepSeek",
        "URL": "https://arxiv.org/abs/2501.12948",
        "발행일": "2025-01",
        "요약": "Rule-based verifier + GRPO의 극단적 성공. 신경망 RM 없이 (a) 정답 문자열 매칭 + (b) <think> 태그 준수 두 규칙만으로 AIME 2024 pass@1 15.6→71.0%(majority vote 86.7%). RLVR 붐의 방아쇠.",
        "관련 섹션": "6. RLVR"
    },
    {
        "번호": 13,
        "제목": "Kimi k1.5: Scaling Reinforcement Learning with LLMs",
        "저자/출처": "Moonshot AI",
        "URL": "https://arxiv.org/abs/2501.12599",
        "발행일": "2025-01",
        "요약": "Rule-based verifier + long-CoT with partial rollout. AIME'24 77.5, MATH-500 96.2. R1과 함께 2025년 초 reasoning model 붐을 대표.",
        "관련 섹션": "6. RLVR"
    },
    {
        "번호": 14,
        "제목": "Qwen2.5-Math Technical Report",
        "저자/출처": "Qwen Team (Alibaba)",
        "URL": "https://arxiv.org/abs/2409.12122",
        "발행일": "2024-09",
        "요약": "Reward Model + self-improvement 파이프라인으로 MATH 83.6, GSM8K 95.9. Qwen2.5-Math-PRM 시리즈 공개, 후속 RLVR 연구의 표준 base model.",
        "관련 섹션": "6. RLVR"
    },
    {
        "번호": 15,
        "제목": "Qwen3 Technical Report",
        "저자/출처": "Qwen Team (Alibaba)",
        "URL": "https://arxiv.org/abs/2505.09388",
        "발행일": "2025-05",
        "요약": "Multi-stage Reasoning RL + General RL 파이프라인, thinking mode toggle. 2025 상반기 오픈 reasoning 모델 SOTA 급.",
        "관련 섹션": "6. RLVR"
    },
    {
        "번호": 16,
        "제목": "Understanding R1-Zero-Like Training: A Critical Perspective (Dr. GRPO)",
        "저자/출처": "Liu et al.",
        "URL": "https://arxiv.org/abs/2503.20783",
        "발행일": "2025-03",
        "요약": "R1-Zero류를 비판적으로 재현. Qwen2.5 base가 프롬프트 없이도 강한 reasoning을 보이는 등 pretraining bias가 상당 부분 설명. GRPO의 길이 편향 제거한 Dr. GRPO 제안. Qwen2.5-Math-7B에서 AIME 43.3%.",
        "관련 섹션": "6. RLVR, 8. 논쟁"
    },
    # Debates & critiques
    {
        "번호": 17,
        "제목": "Does Reinforcement Learning Really Incentivize Reasoning Capacity in LLMs Beyond the Base Model?",
        "저자/출처": "Yue et al. (Tsinghua) — NeurIPS 2025",
        "URL": "https://arxiv.org/abs/2504.13837",
        "발행일": "2025-04",
        "요약": "pass@1은 크게 개선되지만 pass@256에서는 base가 RLVR 앞섬. RLVR path는 이미 base sampling distribution에 존재 → 능력 경계 확장 아님. RLVR 비판 라인의 대표.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 18,
        "제목": "Spurious Rewards: Rethinking Training Signals in RLVR",
        "저자/출처": "Shao et al.",
        "URL": "https://arxiv.org/abs/2506.10947",
        "발행일": "2025-06",
        "요약": "Qwen2.5-Math-7B에서 랜덤 리워드로도 MATH-500 +21.4%p. 원인은 GRPO 클리핑 편향이 pretraining의 '코드로 추론' 습관을 증폭. Llama3/OLMo2에는 통하지 않음 → RLVR 개선분이 capability가 아니라 base prior amplification일 수 있음.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 19,
        "제목": "The Invisible Leash: Why RLVR May or May Not Escape Its Origin",
        "저자/출처": "arXiv 2507.14843",
        "URL": "https://arxiv.org/abs/2507.14843",
        "발행일": "2025-07",
        "요약": "RLVR을 support-constrained optimization으로 형식화. 초기 확률 0인 답은 절대 학습되지 않음, 진행할수록 exploration 축소. pass@1↑, pass@1024↓.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 20,
        "제목": "One Token to Fool LLM-as-a-Judge",
        "저자/출처": "Zhao et al.",
        "URL": "https://arxiv.org/abs/2507.08794",
        "발행일": "2025-07",
        "요약": "GPT-o1, Claude-4 등 leading generative RM들이 ':', '.', 'Thought process:' 같은 단일 토큰으로 false positive를 뱉음. RLVR policy가 이 토큰을 학습하면 리워드 급상승 · 정확도 무관. Master-key attack.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 21,
        "제목": "Hidden Costs of RLVR",
        "저자/출처": "arXiv 2509.21882",
        "URL": "https://arxiv.org/abs/2509.21882",
        "발행일": "2025-09",
        "요약": "RLVR 이후 refusal rate 붕괴 + 오답에 대한 자신감 증가 · miscalibration. reasoning 이득 뒤에 숨은 비용 지적한 position paper.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 22,
        "제목": "Reward Hacking Mitigation using Verifiable Composite Rewards",
        "저자/출처": "arXiv 2509.15557",
        "URL": "https://arxiv.org/abs/2509.15557",
        "발행일": "2025-09",
        "요약": "정답 매칭 + 형식 페널티 + 길이 제약 결합해 hacking 표면 축소. 의료 QA에서 hacking 억제 실증. Composite reward 접근의 대표.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 23,
        "제목": "RLVR Implicitly Incentivizes Correct Reasoning in Base LLMs",
        "저자/출처": "Wen et al.",
        "URL": "https://arxiv.org/abs/2506.14245",
        "발행일": "2025-06",
        "요약": "Yue et al. 반박 라인. pass@1뿐 아니라 계산량-매칭 pass@k에서도 RLVR이 실질적 이득. 특정 조건에서 새로운 경로 창발.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 24,
        "제목": "Large Language Models Cannot Self-Correct Reasoning Yet",
        "저자/출처": "Huang et al. (ICLR 2024)",
        "URL": "https://arxiv.org/abs/2310.01798",
        "발행일": "2023-10",
        "요약": "외부 피드백 없는 intrinsic self-correction이 GSM8K/HotpotQA에서 성능을 오히려 저하. 자기검증 불가능성의 원조 실증.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 25,
        "제목": "Self-Correction Bench: Uncovering and Addressing the Self-Correction Blind Spot",
        "저자/출처": "Jiang et al.",
        "URL": "https://arxiv.org/abs/2507.02778",
        "발행일": "2025-07",
        "요약": "Huang et al. 후속. 동일 모델이 남의 오류는 잡지만 자기 오류에서 체계적 실패 — 'self-correction blind spot'을 형식화.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 26,
        "제목": "The Illusion of Thinking (Apple)",
        "저자/출처": "Shojaee et al. (Apple)",
        "URL": "https://arxiv.org/abs/2506.06941",
        "발행일": "2025-06",
        "요약": "Tower of Hanoi, River Crossing 등 planning puzzle에서 LRM이 특정 복잡도 임계값을 넘으면 accuracy collapse. reasoning token 예산이 남아있어도 사고 노력 감소.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 27,
        "제목": "The Illusion of the Illusion of Thinking (rebuttal)",
        "저자/출처": "Lawsen et al.",
        "URL": "https://arxiv.org/abs/2506.09250",
        "발행일": "2025-06",
        "요약": "Apple 논문의 실험 설계 결함(출력 token limit, unsolvable instance 포함) 지적. 그러나 후속 재현 실험에서도 ~8 disks 근처 실패는 확인.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 28,
        "제목": "Self-Preference Bias in LLM-as-a-Judge",
        "저자/출처": "Panickssery et al.",
        "URL": "https://arxiv.org/abs/2410.21819",
        "발행일": "2024-10",
        "요약": "LLM이 자기 생성 텍스트를 통계적으로 유의미하게 선호. Log-likelihood 정량화. self-recognition 능력과 상관 — 큰 모델일수록 편향 강함.",
        "관련 섹션": "8. 논쟁"
    },
    # Weak-to-Strong & robustness
    {
        "번호": 29,
        "제목": "Shrinking the Generation-Verification Gap with Weak Verifiers (Weaver)",
        "저자/출처": "Saad-Falcon et al. (Stanford Hazy Research)",
        "URL": "https://arxiv.org/abs/2506.18203",
        "발행일": "2025-06",
        "요약": "약한 verifier 앙상블로 generation-verification gap 평균 14.5% 축소 (GPQA Diamond 등). Weak-to-Strong Verification의 실용적 실증.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 30,
        "제목": "Multi-Agent Verification",
        "저자/출처": "arXiv 2502.20379",
        "URL": "https://arxiv.org/abs/2502.20379",
        "발행일": "2025-02",
        "요약": "여러 약한 verifier 앙상블이 강한 generator에도 리프트 — weak-to-strong 실증.",
        "관련 섹션": "8. 논쟁"
    },
    {
        "번호": 31,
        "제목": "CompassVerifier: A Unified and Robust Verifier for LLMs Evaluation and Outcome Reward",
        "저자/출처": "arXiv 2508.03686",
        "URL": "https://arxiv.org/abs/2508.03686",
        "발행일": "2025-08",
        "요약": "통합·강건 verifier 지향. Open-S1 OOD dataset에서 rule-based 대비 우세. Adversarial 학습 요소 포함.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 32,
        "제목": "Adversarial Training of Reward Models",
        "저자/출처": "arXiv 2504.06141",
        "URL": "https://arxiv.org/abs/2504.06141",
        "발행일": "2025-04",
        "요약": "Verifier-annotator 불일치 샘플로 적대적 학습. Master-key 프롬프트에 강건한 판사 모델 학습 프레임워크.",
        "관련 섹션": "7. Verifier Hacking"
    },
    {
        "번호": 33,
        "제목": "From Accuracy to Robustness: A Study of Rule- and Model-based Verifiers",
        "저자/출처": "Huang et al.",
        "URL": "https://arxiv.org/abs/2505.22203",
        "발행일": "2025-05",
        "요약": "Model-based verifier가 rule-based보다 정확도 높지만 정책 최적화 과정에서 지속적으로 hacking돼 인위적 보상 팽창 발생 실증.",
        "관련 섹션": "7. Verifier Hacking"
    },
    # Verifier evaluation benchmarks
    {
        "번호": 34,
        "제목": "ProcessBench: Identifying Process Errors in Mathematical Reasoning",
        "저자/출처": "Zheng et al. (Qwen Team) — ACL 2025",
        "URL": "https://arxiv.org/abs/2412.06559",
        "발행일": "2024-12",
        "요약": "3,400 math CoT에서 첫 오류 step 판정. Human-labeled. Metric은 F1(harmonic of err-acc & correct-acc). 현재 리더보드 최상위 ACTPRM-X 0.760.",
        "관련 섹션": "9. 벤치마크"
    },
    {
        "번호": 35,
        "제목": "PRMBench: A Fine-grained and Challenging Benchmark for Process-Level Reward Models",
        "저자/출처": "Song et al. — ACL 2025",
        "URL": "https://arxiv.org/abs/2501.03124",
        "발행일": "2025-01",
        "요약": "6,216 문제 · 83,456 step 라벨. Simplicity/Soundness/Sensitivity 3축 · 9세부항목. Redundancy, circular logic, deception 등 세밀한 error taxonomy.",
        "관련 섹션": "9. 벤치마크"
    },
    {
        "번호": 36,
        "제목": "MR-Ben: A Meta-Reasoning Benchmark for Evaluating System-2 Thinking in LLMs",
        "저자/출처": "Zeng et al. — NeurIPS 2024",
        "URL": "https://arxiv.org/abs/2406.13975",
        "발행일": "2024-06",
        "요약": "5,975 multi-subject 문제 (physics/chemistry/logic/coding). Model이 solution의 오류 위치 지목 + correction. 수학에 편중된 ProcessBench의 보완.",
        "관련 섹션": "9. 벤치마크"
    },
    {
        "번호": 37,
        "제목": "BIG-Bench Extra Hard (BBEH)",
        "저자/출처": "Kazemi et al. (DeepMind) — ACL 2025",
        "URL": "https://arxiv.org/abs/2502.19187",
        "발행일": "2025-02",
        "요약": "BBH의 23개 태스크를 어려운 버전으로 교체 · 4,520 examples. Finding errors in reasoning traces가 명시적 skill로 포함. Best reasoning 54.2%, best general 23.9%.",
        "관련 섹션": "9. 벤치마크"
    },
    {
        "번호": 38,
        "제목": "REVEAL — A Chain-of-Thought Is as Strong as Its Weakest Link",
        "저자/출처": "arXiv 2402.00559",
        "URL": "https://arxiv.org/abs/2402.00559",
        "발행일": "2024-02",
        "요약": "Reasoning chain step verification 초기 벤치마크. GPT-4 baseline ~0.65 range.",
        "관련 섹션": "9. 벤치마크"
    },
    {
        "번호": 39,
        "제목": "Qwen2.5-Math-PRM: Blog & Model",
        "저자/출처": "Qwen Team",
        "URL": "https://qwenlm.github.io/blog/qwen2.5-math-prm/",
        "발행일": "2025",
        "요약": "Qwen2.5-Math-PRM-7B (ProcessBench 0.75 급) 및 학습 파이프라인 공개. 오픈소스 SOTA PRM 중 하나.",
        "관련 섹션": "9. 벤치마크"
    },
    # Community references
    {
        "번호": 40,
        "제목": "awesome-RLVR",
        "저자/출처": "OpenDILab (GitHub)",
        "URL": "https://github.com/opendilab/awesome-RLVR",
        "발행일": "2025",
        "요약": "RLVR 논문 · 코드 · 데이터셋 큐레이션 리스트. 이 리서치에서 다룬 논문 대부분을 포함.",
        "관련 섹션": "10. 참고문헌"
    },
]


def create_references_xlsx():
    """참고문헌 XLSX 파일을 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ROW_FILL_1 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    ROW_FILL_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    HEADER_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Malgun Gothic", size=10)
    URL_FONT = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    BOLD_FONT = Font(name="Malgun Gothic", bold=True, size=10)

    THIN_BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 55, 30, 45, 10, 65, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, ref in enumerate(references, 2):
        is_odd = (row_idx % 2 == 0)
        row_fill = ROW_FILL_1 if is_odd else ROW_FILL_2

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER

            if col_idx == 1:
                cell.font = BOLD_FONT
                cell.alignment = CENTER
            elif col_idx == 4:
                cell.font = URL_FONT
                cell.alignment = LEFT
                cell.hyperlink = value
            else:
                cell.font = BODY_FONT
                cell.alignment = LEFT

        ws.row_dimensions[row_idx].height = 70

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    output_path = "llm-as-verifier-references.xlsx"
    wb.save(output_path)
    print(f"[완료] 참고문헌 파일 생성: {output_path}")
    print(f"[정보] 총 {len(references)}개 참고문헌 포함")
    return output_path


if __name__ == "__main__":
    create_references_xlsx()
