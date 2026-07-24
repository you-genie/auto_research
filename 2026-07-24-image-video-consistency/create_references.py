"""
참고문헌 XLSX 파일 생성 스크립트
이미지·비디오 생성의 Consistency 완전 정복 (2022~2026)
"""

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

REFERENCES = [
    # 이미지 — 최적화 시대
    {
        "번호": 1,
        "제목": "An Image is Worth One Word: Personalizing Text-to-Image Generation using Textual Inversion",
        "저자/출처": "Gal et al. (Tel Aviv U / NVIDIA)",
        "URL": "https://arxiv.org/abs/2208.01618",
        "발행일": "2022-08 (ICLR'23)",
        "요약 (한국어)": "모델을 동결하고 개념을 재현하는 새 텍스트 임베딩 벡터 하나(pseudo-word)만 학습하는 개인화 기법. 용량은 극소형이나 복잡한 정체성 충실도가 낮음.",
        "관련 섹션": "① 최적화 시대",
    },
    {
        "번호": 2,
        "제목": "DreamBooth: Fine Tuning Text-to-Image Diffusion Models for Subject-Driven Generation",
        "저자/출처": "Ruiz et al. (Google)",
        "URL": "https://arxiv.org/abs/2208.12242",
        "발행일": "2022-08 (CVPR'23)",
        "요약 (한국어)": "희귀 토큰에 피사체를 묶고 UNet 전체를 미세조정. prior-preservation loss로 언어 드리프트·과적합 방지. 고충실도지만 피사체당 수 GB 체크포인트.",
        "관련 섹션": "① 최적화 시대",
    },
    {
        "번호": 3,
        "제목": "LoRA: Low-Rank Adaptation of Large Language Models",
        "저자/출처": "Hu et al. (Microsoft)",
        "URL": "https://arxiv.org/abs/2106.09685",
        "발행일": "2021-06",
        "요약 (한국어)": "가중치를 저랭크 업데이트(BA)로만 갱신하는 경량 미세조정. 커뮤니티가 디퓨전에 이식해 오늘날 SDXL/FLUX 커스터마이징의 사실상 표준 포맷.",
        "관련 섹션": "① 최적화 시대",
    },
    {
        "번호": 4,
        "제목": "Multi-Concept Customization of Text-to-Image Diffusion (Custom Diffusion)",
        "저자/출처": "Kumari et al. (CMU / Adobe)",
        "URL": "https://arxiv.org/abs/2212.04488",
        "발행일": "2022-12 (CVPR'23)",
        "요약 (한국어)": "cross-attention의 K/V 투영만 미세조정해도 충분함을 보임. '정체성은 cross-attn K/V에 산다'는 통찰로 이후 인코더 기법의 사상적 뿌리.",
        "관련 섹션": "① 최적화 시대",
    },
    # 이미지 — 인코더 시대
    {
        "번호": 5,
        "제목": "IP-Adapter: Text Compatible Image Prompt Adapter for Text-to-Image Diffusion Models",
        "저자/출처": "Ye et al. (Tencent)",
        "URL": "https://arxiv.org/abs/2308.06721",
        "발행일": "2023-08",
        "요약 (한국어)": "Decoupled cross-attention — CLIP 이미지 토큰용 별도 K/V 어텐션을 병렬 추가해 텍스트와 충돌 없이 이미지 프롬프트 주입. 얼굴 신원은 중간.",
        "관련 섹션": "② 인코더 시대",
    },
    {
        "번호": 6,
        "제목": "PhotoMaker: Customizing Realistic Human Photos via Stacked ID Embedding",
        "저자/출처": "Li et al. (Tencent ARC / Nankai)",
        "URL": "https://arxiv.org/abs/2312.04461",
        "발행일": "2023-12 (CVPR'24)",
        "요약 (한국어)": "같은 사람의 여러 장을 인코딩해 하나의 통합 ID로 stack/pool. 편집성·다양성 우수, 정체성 블렌딩 가능. 사람 전용.",
        "관련 섹션": "② 인코더 시대",
    },
    {
        "번호": 7,
        "제목": "InstantID: Zero-shot Identity-Preserving Generation in Seconds",
        "저자/출처": "Wang et al. (InstantX)",
        "URL": "https://arxiv.org/abs/2401.07519",
        "발행일": "2024-01",
        "요약 (한국어)": "ArcFace ID 임베딩(의미) + 5점 랜드마크 IdentityNet(공간)으로 1장 고충실도. 단 랜드마크가 포즈·표정을 과하게 고정.",
        "관련 섹션": "② 인코더 시대",
    },
    {
        "번호": 8,
        "제목": "PuLID: Pure and Lightning ID Customization via Contrastive Alignment",
        "저자/출처": "Guo et al. (ByteDance)",
        "URL": "https://arxiv.org/abs/2404.16022",
        "발행일": "2024-04 (NeurIPS'24)",
        "요약 (한국어)": "Lightning 분기로 contrastive alignment loss + accurate ID loss 구현. 신원 충실도+편집성 SOTA, 배경·조명 침습 최소. PuLID-FLUX 존재.",
        "관련 섹션": "② 인코더 시대",
    },
    # 이미지 — 학습-프리 세트 & 스타일
    {
        "번호": 9,
        "제목": "The Chosen One: Consistent Characters in Text-to-Image Diffusion Models",
        "저자/출처": "Avrahami et al. (Google)",
        "URL": "https://arxiv.org/abs/2311.10093",
        "발행일": "2023-11 (SIGGRAPH'24)",
        "요약 (한국어)": "갤러리 생성 → 임베딩 군집화 → 응집 클러스터 선택 → 개인화를 반복해 정체성 수렴. 무학습은 아니고 반복적, 얻는 정체성이 창발적.",
        "관련 섹션": "③ 학습-프리 세트 일관성",
    },
    {
        "번호": 10,
        "제목": "Training-Free Consistent Text-to-Image Generation (ConsiStory)",
        "저자/출처": "Tewel et al. (NVIDIA / Tel Aviv U)",
        "URL": "https://arxiv.org/abs/2402.03286",
        "발행일": "2024-02 (SIGGRAPH'24)",
        "요약 (한국어)": "Subject-Driven Self-Attention — 배치 내 다른 이미지의 피사체 패치 K/V까지 공유(마스킹) + DIFT 특징 주입. H100 ~10초/장, 무학습.",
        "관련 섹션": "③ 학습-프리 세트 일관성",
    },
    {
        "번호": 11,
        "제목": "StoryDiffusion: Consistent Self-Attention for Long-Range Image and Video Generation",
        "저자/출처": "Zhou et al. (Nankai / ByteDance)",
        "URL": "https://arxiv.org/abs/2405.01434",
        "발행일": "2024-05 (NeurIPS'24)",
        "요약 (한국어)": "self-attention을 대체하는 드롭인 Consistent Self-Attention — 배치 토큰을 K/V에 이어붙여 외형 수렴. Semantic Motion Predictor로 영상 확장.",
        "관련 섹션": "③ 학습-프리 세트 일관성",
    },
    {
        "번호": 12,
        "제목": "One-Prompt-One-Story: Free-Lunch Consistent T2I Generation Using a Single Prompt",
        "저자/출처": "Liu et al.",
        "URL": "https://arxiv.org/abs/2501.13554",
        "발행일": "2025-01 (ICLR'25 Spotlight)",
        "요약 (한국어)": "'context consistency' 통찰 — 모든 프레임 프롬프트를 하나로 연결해 정체성 고정. SVR + Identity-Preserving Cross-Attention. 프레임 수가 프롬프트 길이에 제약.",
        "관련 섹션": "③ 학습-프리 세트 일관성",
    },
    {
        "번호": 13,
        "제목": "Style Aligned Image Generation via Shared Attention (StyleAligned)",
        "저자/출처": "Hertz et al. (Google)",
        "URL": "https://arxiv.org/abs/2312.02133",
        "발행일": "2023-12 (CVPR'24)",
        "요약 (한국어)": "세트를 함께 생성하며 self-attention이 레퍼런스를 참조하되 Q·K를 AdaIN 정규화해 균형 잡힌 스타일 전이. 실사엔 DDIM inversion 병용.",
        "관련 섹션": "④ 스타일 일관성",
    },
    {
        "번호": 14,
        "제목": "InstantStyle: Free Lunch towards Style-Preserving in Text-to-Image Generation",
        "저자/출처": "Wang et al. (InstantX / Xiaohongshu)",
        "URL": "https://arxiv.org/abs/2404.02733",
        "발행일": "2024-04",
        "요약 (한국어)": "CLIP 공간에서 이미지-내용 임베딩을 빼 순수 스타일 추출 + 스타일 블록에만 주입해 내용 누출 방지. 무학습, SDXL 특화 블록.",
        "관련 섹션": "④ 스타일 일관성",
    },
    {
        "번호": 15,
        "제목": "Implicit Style-Content Separation using B-LoRA",
        "저자/출처": "Frenkel et al. (Reichman / Tel Aviv U)",
        "URL": "https://arxiv.org/abs/2403.14572",
        "발행일": "2024-03",
        "요약 (한국어)": "SDXL의 특정 두 블록에 LoRA를 함께 학습하면 1장으로 스타일 vs 내용이 자연 분리. 두 B-LoRA 재조합으로 스타일 전이. SDXL 특화.",
        "관련 섹션": "④ 스타일 일관성",
    },
    # 이미지 — 네이티브·통합
    {
        "번호": 16,
        "제목": "OmniGen: Unified Image Generation (및 OmniGen2)",
        "저자/출처": "Xiao et al. (BAAI)",
        "URL": "https://arxiv.org/abs/2409.11340",
        "발행일": "2024-09 / 2025 (OmniGen2: 2506.18871)",
        "요약 (한국어)": "단일 디퓨전 트랜스포머가 T2I·편집·피사체 기반 생성을 인터리브 지시로 처리 — ControlNet/IP-Adapter 불필요. 통합 in-context 조건에서 일관성 창발.",
        "관련 섹션": "⑤ 네이티브·통합 시대",
    },
    {
        "번호": 17,
        "제목": "Transfusion: Predict the Next Token and Diffuse Images with One Multi-Modal Model",
        "저자/출처": "Zhou et al. (Meta)",
        "URL": "https://arxiv.org/abs/2408.11039",
        "발행일": "2024-08",
        "요약 (한국어)": "자기회귀 토큰 예측 + 디퓨전을 단일 모델에 통합. GPT-4o 네이티브 이미지 생성(tokens→transformer→diffusion→pixels)의 개념적 배경.",
        "관련 섹션": "⑤ 네이티브·통합 시대",
    },
    # 비디오 — 시간 기초 & 캐릭터
    {
        "번호": 18,
        "제목": "Video Diffusion Models",
        "저자/출처": "Ho et al. (Google)",
        "URL": "https://arxiv.org/abs/2204.03458",
        "발행일": "2022-04",
        "요약 (한국어)": "2D UNet을 시공간 3D로 확장하되 분해(factorized) — 공간 블록 뒤 temporal attention 삽입. 이미지·비디오 공동 학습. 저해상도·짧은 클립.",
        "관련 섹션": "⑥ 시간적 일관성 기초",
    },
    {
        "번호": 19,
        "제목": "Align your Latents: High-Resolution Video Synthesis with Latent Diffusion Models (Video LDM)",
        "저자/출처": "Blattmann et al. (NVIDIA)",
        "URL": "https://arxiv.org/abs/2304.08818",
        "발행일": "2023-04 (CVPR'23)",
        "요약 (한국어)": "사전학습 SD의 공간 가중치를 동결하고 시간 층(3D conv + temporal attention)만 비디오로 학습. 다른 체크포인트에 일반화, 고해상도 가능.",
        "관련 섹션": "⑥ 시간적 일관성 기초",
    },
    {
        "번호": 20,
        "제목": "AnimateDiff: Animate Your Personalized T2I Diffusion Models without Specific Tuning",
        "저자/출처": "Guo et al.",
        "URL": "https://arxiv.org/abs/2307.04725",
        "발행일": "2023-07 (ICLR'24 Spotlight)",
        "요약 (한국어)": "분리·이식 가능한 모션 모듈(사인 위치인코딩 temporal self-attention)을 한 번 학습해 임의의 개인화 체크포인트에 꽂아 애니메이션. ~16프레임 창.",
        "관련 섹션": "⑥ 시간적 일관성 기초",
    },
    {
        "번호": 21,
        "제목": "FreeInit: Bridging Initialization Gap in Video Diffusion Models",
        "저자/출처": "Wu et al.",
        "URL": "https://arxiv.org/abs/2312.07537",
        "발행일": "2023-12 (ECCV'24)",
        "요약 (한국어)": "추론 노이즈의 저주파 시공간 성분이 학습 때와 다른 간극을 발견, 반복 재샘플·정제로 재학습 없이 시간 일관성 향상. 추론 비용 증가.",
        "관련 섹션": "⑥ 시간적 일관성 기초",
    },
    {
        "번호": 22,
        "제목": "Animate Anyone: Consistent and Controllable Image-to-Video Synthesis for Character Animation",
        "저자/출처": "Hu et al. (Alibaba)",
        "URL": "https://arxiv.org/abs/2311.17117",
        "발행일": "2023-11 (CVPR'24)",
        "요약 (한국어)": "ReferenceNet(UNet 복사본)이 레퍼런스를 인코딩해 self-attention K/V로 융합, 정교한 텍스처 전이. Pose Guider + temporal 층. 메모리 2배.",
        "관련 섹션": "⑦ 캐릭터 애니메이션",
    },
    {
        "번호": 23,
        "제목": "MagicAnimate: Temporally Consistent Human Image Animation using Diffusion Model",
        "저자/출처": "Xu et al. (NUS Show Lab / ByteDance)",
        "URL": "https://arxiv.org/abs/2311.16498",
        "발행일": "2023-11 (CVPR'24)",
        "요약 (한국어)": "전용 appearance encoder + temporal attention + DensePose(조밀 표면) 조건으로 매끄러운 모션. 1인 전용, 윈도우 경계 흐림.",
        "관련 섹션": "⑦ 캐릭터 애니메이션",
    },
    {
        "번호": 24,
        "제목": "Champ: Controllable and Consistent Human Image Animation with 3D Parametric Guidance",
        "저자/출처": "Zhu et al.",
        "URL": "https://arxiv.org/abs/2403.14781",
        "발행일": "2024-03 (ECCV'24)",
        "요약 (한국어)": "2D 포즈 대신 SMPL 3D 메시 → 깊이·법선·시맨틱·스켈레톤 다층 조건. 3D 사전이 형상·포즈 일관성과 교차 신원 전이 개선.",
        "관련 섹션": "⑦ 캐릭터 애니메이션",
    },
    # 비디오 — 파운데이션 & DiT
    {
        "번호": 25,
        "제목": "Stable Video Diffusion: Scaling Latent Video Diffusion Models to Large Datasets",
        "저자/출처": "Blattmann et al. (Stability AI)",
        "URL": "https://arxiv.org/abs/2311.15127",
        "발행일": "2023-11",
        "요약 (한국어)": "이미지 LDM + 시간 층, 3단계 커리큘럼(이미지→비디오 사전학습→고품질 미세조정). 데이터 큐레이션이 시간 품질의 최대 지렛대임을 입증.",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    {
        "번호": 26,
        "제목": "Scalable Diffusion Models with Transformers (DiT)",
        "저자/출처": "Peebles & Xie (UC Berkeley)",
        "URL": "https://arxiv.org/abs/2212.09748",
        "발행일": "2022-12",
        "요약 (한국어)": "UNet 대신 트랜스포머를 디퓨전 백본으로. Sora·CogVideoX 등 시공간 패치 토큰 + 풀 어텐션 비디오 모델의 아키텍처 기반.",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    {
        "번호": 27,
        "제목": "Sora: Video generation models as world simulators (Technical Report)",
        "저자/출처": "OpenAI",
        "URL": "https://openai.com/index/video-generation-models-as-world-simulators/",
        "발행일": "2024-02",
        "요약 (한국어)": "비디오를 시공간 잠재 패치 토큰으로 만들어 트랜스포머로 학습. 모든 토큰에 전역 어텐션 → 장기 시간 일관성·물체 지속성이 스케일에서 창발. 최대 ~1분.",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    {
        "번호": 28,
        "제목": "CogVideoX: Text-to-Video Diffusion Models with An Expert Transformer",
        "저자/출처": "Yang et al. (Zhipu / Tsinghua)",
        "URL": "https://arxiv.org/abs/2408.06072",
        "발행일": "2024-08",
        "요약 (한국어)": "3D causal VAE + 3D 풀 어텐션 expert Transformer로 프레임 간 불일치를 겨냥해 분해형 어텐션을 대체. 대표 오픈 DiT 비디오 모델.",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    {
        "번호": 29,
        "제목": "HunyuanVideo: A Systematic Framework For Large Video Generative Models",
        "저자/출처": "Kong et al. (Tencent)",
        "URL": "https://arxiv.org/abs/2412.03603",
        "발행일": "2024-12",
        "요약 (한국어)": "13B 파라미터, causal 3D VAE, MMDiT류 풀 어텐션 트랜스포머, LLM 텍스트 인코더. 대규모 오픈 비디오 생성 프레임워크.",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    {
        "번호": 30,
        "제목": "Wan: Open and Advanced Large-Scale Video Generative Models",
        "저자/출처": "Alibaba (Wan Team)",
        "URL": "https://arxiv.org/abs/2503.20314",
        "발행일": "2025-03",
        "요약 (한국어)": "효율적 3D causal Wan-VAE + DiT. 이미지+비디오를 아우르는 강력한 오픈 베이스라인(Wan 2.1/2.2).",
        "관련 섹션": "⑧ 파운데이션 & DiT",
    },
    # 비디오 — 긴 영상·멀티샷·월드
    {
        "번호": 31,
        "제목": "StreamingT2V: Consistent, Dynamic, and Extendable Long Video Generation from Text",
        "저자/출처": "Henschel et al. (Picsart)",
        "URL": "https://arxiv.org/abs/2403.14773",
        "발행일": "2024-03 (CVPR'25)",
        "요약 (한국어)": "CAM(단기 메모리 주입) + APM(앵커로 장기 정체성)으로 청크 단위 자기회귀 생성. 200+ 프레임. 오차 누적·모션 균질화가 한계.",
        "관련 섹션": "⑨ 긴 영상·멀티샷",
    },
    {
        "번호": 32,
        "제목": "FreeNoise: Tuning-Free Longer Video Diffusion via Noise Rescheduling",
        "저자/출처": "Qiu et al.",
        "URL": "https://arxiv.org/abs/2310.15169",
        "발행일": "2023-10 (ICLR'24)",
        "요약 (한국어)": "노이즈를 창 단위로 셔플·재사용해 장거리 상관 유지 + 윈도우 temporal attention 융합. +17% 시간으로 긴 클립. 내용이 초기 프롬프트를 반복.",
        "관련 섹션": "⑨ 긴 영상·멀티샷",
    },
    {
        "번호": 33,
        "제목": "Gen-L-Video: Multi-Text to Long Video Generation via Temporal Co-Denoising",
        "저자/출처": "Wang et al.",
        "URL": "https://arxiv.org/abs/2305.18264",
        "발행일": "2023-05",
        "요약 (한국어)": "긴 비디오를 겹치는 짧은 구간으로 나눠 병렬 디노이즈 후 겹침 평균. 오프더셸프 짧은 모델로 임의 길이. 평균화로 모션 흐림.",
        "관련 섹션": "⑨ 긴 영상·멀티샷",
    },
    {
        "번호": 34,
        "제목": "VideoStudio: Generating Consistent-Content and Multi-Scene Videos",
        "저자/출처": "Long et al.",
        "URL": "https://arxiv.org/abs/2401.01256",
        "발행일": "2024-01 (ECCV'24)",
        "요약 (한국어)": "LLM이 프롬프트를 다장면 스크립트로, 공통 엔티티별 레퍼런스 이미지 생성 후 각 장면을 그 레퍼런스로 조건화해 교차 장면 일관성 확보.",
        "관련 섹션": "⑨ 긴 영상·멀티샷",
    },
    {
        "번호": 35,
        "제목": "MotionCtrl: A Unified and Flexible Motion Controller for Video Generation",
        "저자/출처": "Wang et al. (Tencent ARC)",
        "URL": "https://arxiv.org/abs/2312.03641",
        "발행일": "2023-12 (SIGGRAPH'24)",
        "요약 (한국어)": "CMCM(카메라 외부파라미터)과 OMCM(객체 궤적)을 분리 학습해 카메라·객체 모션을 독립 제어. 거친 제어, 궤적 주석 필요.",
        "관련 섹션": "⑩ 카메라·3D·월드",
    },
    {
        "번호": 36,
        "제목": "CameraCtrl: Enabling Camera Control for Text-to-Video Generation",
        "저자/출처": "He et al.",
        "URL": "https://arxiv.org/abs/2404.02101",
        "발행일": "2024-04",
        "요약 (한국어)": "픽셀별 Plücker ray 임베딩(카메라 포즈의 기하 표현)을 카메라 인코더가 받아 temporal attention에 주입. 원시 외부파라미터보다 일반화 우수. 카메라 전용.",
        "관련 섹션": "⑩ 카메라·3D·월드",
    },
    {
        "번호": 37,
        "제목": "Genie: Generative Interactive Environments",
        "저자/출처": "Bruce et al. (Google DeepMind)",
        "URL": "https://arxiv.org/abs/2402.15391",
        "발행일": "2024-02 (ICML'24 Best Paper)",
        "요약 (한국어)": "시공간 토크나이저 + 잠재 행동 모델(비지도) + 자기회귀 동역학 모델. 11B. Genie 3(2025)는 환경 기억 ~1분 유지, 창발적 공간 지속성.",
        "관련 섹션": "⑩ 카메라·3D·월드",
    },
    # 샘플링 — Consistency Models (일관성 아님)
    {
        "번호": 38,
        "제목": "Consistency Models",
        "저자/출처": "Song, Dhariwal, Chen, Sutskever (OpenAI)",
        "URL": "https://arxiv.org/abs/2303.01469",
        "발행일": "2023-03 (ICML'23)",
        "요약 (한국어)": "⚠️ 시간·캐릭터 일관성이 아님! PF-ODE 궤적의 임의 점을 시작점으로 매핑하는 자기 일관성으로 1스텝 생성. 빠른 샘플링/증류 기법.",
        "관련 섹션": "⑪ 함정: Consistency Models",
    },
    {
        "번호": 39,
        "제목": "Latent Consistency Models: Synthesizing High-Resolution Images with Few-Step Inference",
        "저자/출처": "Luo et al. (Tsinghua)",
        "URL": "https://arxiv.org/abs/2310.04378",
        "발행일": "2023-10",
        "요약 (한국어)": "CM을 잠재공간(SD)에 적용, CFG 통합 증강 PF-ODE로 ~1~4 스텝 고해상도. ~32 A100-시간에 증류. 목적은 속도지 시각 일관성이 아님.",
        "관련 섹션": "⑪ 함정: Consistency Models",
    },
    {
        "번호": 40,
        "제목": "LCM-LoRA: A Universal Stable-Diffusion Acceleration Module",
        "저자/출처": "Luo et al.",
        "URL": "https://arxiv.org/abs/2311.05556",
        "발행일": "2023-11",
        "요약 (한국어)": "가속을 LoRA 모듈로 증류한 범용 가속 모듈. 여러 SD 체크포인트에 이식 가능. AnimateLCM 등으로 비디오 가속에도 사용.",
        "관련 섹션": "⑪ 함정: Consistency Models",
    },
]


def create_xlsx(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 스타일 정의 (딥 퍼플/시네마 테마)
    header_fill = PatternFill(start_color="6C3EF5", end_color="6C3EF5", fill_type="solid")
    even_fill = PatternFill(start_color="ECE6FD", end_color="ECE6FD", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Malgun Gothic", size=10)
    link_font = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 행
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 55, 30, 45, 18, 62, 26]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, ref in enumerate(REFERENCES, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약 (한국어)"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            ws.row_dimensions[row_idx].height = 66

            if col_idx == 1:  # 번호
                cell.alignment = center_align
                cell.font = Font(name="Malgun Gothic", size=10, bold=True)
            elif col_idx == 4:  # URL — 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = link_font
                cell.alignment = left_align
            elif col_idx == 5:  # 발행일
                cell.alignment = center_align
                cell.font = body_font
            else:
                cell.alignment = left_align
                cell.font = body_font

    # 틀 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "6C3EF5"

    wb.save(output_path)
    print(f"XLSX 파일 생성 완료: {output_path}")
    print(f"총 참고문헌 수: {len(REFERENCES)}개")


if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "image-video-consistency-references.xlsx")
    create_xlsx(output_path)
