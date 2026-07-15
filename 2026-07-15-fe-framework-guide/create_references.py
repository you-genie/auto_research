"""
참고문헌 XLSX 파일 생성 스크립트
프론트엔드 프레임워크 완전 정복 (2025~2026) — 비교 · 개발 주의점 · 코드 가이드
"""

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

REFERENCES = [
    {
        "번호": 1,
        "제목": "State of JavaScript 2025 — Front-end Frameworks",
        "저자/출처": "State of JS (Devographics)",
        "URL": "https://2025.stateofjs.com/en-US/libraries/front-end-frameworks/",
        "발행일": "2026-03-01",
        "요약 (한국어)": "6천여 개발자 설문. 지난 1년간 프론트엔드 사용 순위가 거의 변동 없음(Alpine/HTMX만 교체). Solid가 5년 연속 최고 만족도, 생태계가 '요동침'에서 '안정됨'으로 전환.",
        "관련 섹션": "프레임워크 지형도",
    },
    {
        "번호": 2,
        "제목": "State of JavaScript 2025: A Maturing Ecosystem with TypeScript Cementing Dominance",
        "저자/출처": "InfoQ",
        "URL": "https://www.infoq.com/news/2026/03/state-of-js-survey-2025/",
        "발행일": "2026-03-01",
        "요약 (한국어)": "State of JS 2025 해설. TypeScript 지배력 강화, 응답자 행복도 5년째 3.8/5 유지, 프레임워크 창조 시대의 종료와 생태계 성숙을 분석.",
        "관련 섹션": "프레임워크 지형도 / 코드 가이드",
    },
    {
        "번호": 3,
        "제목": "React 19 vs Vue 3.6 vs Svelte 5: 2026 Framework Convergence",
        "저자/출처": "byteiota",
        "URL": "https://byteiota.com/react-19-vs-vue-3-6-vs-svelte-5-2026-framework-convergence/",
        "발행일": "2026-01-01",
        "요약 (한국어)": "세 프레임워크가 fine-grained reactivity·서버 우선 렌더링·컴파일러 최적화로 수렴 중. 성능 벤치마크(ops/sec)와 번들 크기(Svelte 28KB, Vue 58KB/Vapor 10KB, React 72KB) 비교.",
        "관련 섹션": "프레임워크별 상세 비교",
    },
    {
        "번호": 4,
        "제목": "React vs Vue vs Angular vs Svelte: Framework Performance Reality 2025",
        "저자/출처": "JavaScript in Plain English",
        "URL": "https://javascript.plainenglish.io/react-vs-vue-vs-angular-vs-svelte-framework-performance-reality-2025-52f1414cf0b8",
        "발행일": "2025-08-01",
        "요약 (한국어)": "4대 프레임워크의 실제 성능 현실을 벤치마크로 비교. 측정 조건에 따른 편차와 '점유율 vs 성능'의 괴리를 지적.",
        "관련 섹션": "프레임워크별 상세 비교",
    },
    {
        "번호": 5,
        "제목": "JavaScript Framework Trends in 2026: React, Next.js, Vue, Angular, Svelte",
        "저자/출처": "Nucamp",
        "URL": "https://www.nucamp.co/blog/javascript-framework-trends-in-2026-what-s-new-in-react-next.js-vue-angular-and-svelte",
        "발행일": "2026-01-01",
        "요약 (한국어)": "2026년 기준 각 프레임워크의 최신 변화 요약. React Compiler·Server Components, Vue Vapor Mode, Angular Signals+Zoneless, Svelte Runes.",
        "관련 섹션": "프레임워크별 상세 비교",
    },
    {
        "번호": 6,
        "제목": "React 19 vs Svelte 5 vs Vue 4 Performance Benchmarks 2025",
        "저자/출처": "jsgurujobs",
        "URL": "https://jsgurujobs.com/blog/svelte-5-vs-react-19-vs-vue-4-the-2025-framework-war-nobody-expected-performance-benchmarks",
        "발행일": "2025-09-01",
        "요약 (한국어)": "Lighthouse 점수와 번들 크기 중심 벤치마크. Svelte 5가 성능·번들에서 앞서지만 React가 시장 점유율 44.7%로 지배함을 대조.",
        "관련 섹션": "프레임워크별 상세 비교",
    },
    {
        "번호": 7,
        "제목": "React 19 Documentation",
        "저자/출처": "Meta / React Team",
        "URL": "https://react.dev/blog/2024/12/05/react-19",
        "발행일": "2024-12-05",
        "요약 (한국어)": "React 19 공식 릴리스 노트. Actions, useActionState, useOptimistic, use() 훅, Server Components 안정화, ref as prop 등 핵심 변경사항.",
        "관련 섹션": "React 19",
    },
    {
        "번호": 8,
        "제목": "React Compiler",
        "저자/출처": "React Team",
        "URL": "https://react.dev/learn/react-compiler",
        "발행일": "2025-04-01",
        "요약 (한국어)": "useMemo/useCallback/memo를 자동화하는 React Compiler 공식 문서. 컴파일 시점 메모이제이션으로 불필요한 리렌더를 제거하는 원리와 도입 방법.",
        "관련 섹션": "React 19 - Compiler",
    },
    {
        "번호": 9,
        "제목": "Vue 3.5 Release Notes",
        "저자/출처": "Vue.js Team (Evan You)",
        "URL": "https://blog.vuejs.org/posts/vue-3-5",
        "발행일": "2024-09-01",
        "요약 (한국어)": "Vue 3.5 공식 발표. Reactive Props Destructure, useId, useTemplateRef, 메모리 최적화 및 Vapor Mode(가상 DOM 우회) 방향성 소개.",
        "관련 섹션": "Vue 3.5",
    },
    {
        "번호": 10,
        "제목": "Angular Signals & Zoneless Guide",
        "저자/출처": "Angular Team (Google)",
        "URL": "https://angular.dev/guide/signals",
        "발행일": "2025-05-01",
        "요약 (한국어)": "Angular Signals 공식 가이드. signal()/computed()/effect() 기반 세밀 반응성과 Zoneless 변경 감지, Standalone Components로의 전환.",
        "관련 섹션": "Angular 20",
    },
    {
        "번호": 11,
        "제목": "Svelte 5 Runes",
        "저자/출처": "Svelte Team",
        "URL": "https://svelte.dev/docs/svelte/what-are-runes",
        "발행일": "2024-10-01",
        "요약 (한국어)": "Svelte 5 Runes($state/$derived/$effect) 공식 문서. 암묵적 최상위 반응성을 명시적 시그널로 대체해 컴포넌트 밖에서도 동작하게 만든 새 반응성 모델.",
        "관련 섹션": "Svelte 5",
    },
    {
        "번호": 12,
        "제목": "SolidJS Documentation",
        "저자/출처": "Solid Team (Ryan Carniato)",
        "URL": "https://www.solidjs.com/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "진짜 fine-grained signals 기반 프레임워크. JSX를 쓰되 컴포넌트가 한 번만 실행되고 시그널이 DOM을 직접 갱신. State of JS 5년 연속 최고 만족도.",
        "관련 섹션": "Solid & Qwik",
    },
    {
        "번호": 13,
        "제목": "Qwik — Resumable Framework",
        "저자/출처": "Builder.io / Qwik Team",
        "URL": "https://qwik.dev/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Resumability(재개성) 아키텍처로 하이드레이션을 제거해 초기 상호작용까지의 JS를 거의 0으로 만드는 프레임워크. 대규모 콘텐츠 사이트에 유리.",
        "관련 섹션": "Solid & Qwik",
    },
    {
        "번호": 14,
        "제목": "The 2025 Frontend Framework Showdown: Next.js, Nuxt.js, SvelteKit, Astro",
        "저자/출처": "Leapcell",
        "URL": "https://leapcell.io/blog/the-2025-frontend-framework-showdown-next-js-nuxt-js-sveltekit-and-astro",
        "발행일": "2025-06-01",
        "요약 (한국어)": "메타프레임워크 4종 비교. 렌더링 전략(SSR/SSG/ISR), 생태계, 성능, 적합 사용처를 표로 정리. 블로그·문서는 Astro, Vue 대규모는 Nuxt 등 추천.",
        "관련 섹션": "메타프레임워크와 렌더링 전략",
    },
    {
        "번호": 15,
        "제목": "Next.js vs Remix vs Astro vs SvelteKit in 2026: Definitive Decision Guide",
        "저자/출처": "DEV Community (Pockit)",
        "URL": "https://dev.to/pockit_tools/nextjs-vs-remix-vs-astro-vs-sveltekit-in-2026-the-definitive-framework-decision-guide-lp5",
        "발행일": "2026-02-01",
        "요약 (한국어)": "Next.js 16 PPR/Cache Components, Astro의 Cloudflare 인수(2026-01), Remix의 React Router v7 합류, SvelteKit의 JS 절감 등 2026 메타프레임워크 지형 정리.",
        "관련 섹션": "메타프레임워크와 렌더링 전략",
    },
    {
        "번호": 16,
        "제목": "Next.js Documentation (App Router, PPR)",
        "저자/출처": "Vercel",
        "URL": "https://nextjs.org/docs",
        "발행일": "2025-10-01",
        "요약 (한국어)": "Next.js 공식 문서. App Router, Server Components, Partial Prerendering(정적 셸 즉시 + 동적 스트리밍), Cache Components 등 렌더링 전략.",
        "관련 섹션": "메타프레임워크 - Next.js",
    },
    {
        "번호": 17,
        "제목": "Frontend Performance Checklist For 2025",
        "저자/출처": "Crystallize",
        "URL": "https://crystallize.com/blog/frontend-performance-checklist",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Core Web Vitals 목표(INP ≤200ms, LCP ≤2.5s, CLS ≤0.1), 성능 예산(JS ≤400KB gzip), 하이드레이션 최소화, RUM 모니터링 등 실무 성능 체크리스트.",
        "관련 섹션": "개발 주의점 - 성능",
    },
    {
        "번호": 18,
        "제목": "Web Vitals (LCP, INP, CLS)",
        "저자/출처": "Google web.dev",
        "URL": "https://web.dev/articles/vitals",
        "발행일": "2024-03-01",
        "요약 (한국어)": "Core Web Vitals 공식 가이드. 2024년 3월 INP가 FID를 대체. 각 지표의 정의·측정·개선 방법과 필드(RUM) 데이터 수집 방식.",
        "관련 섹션": "개발 주의점 - 성능",
    },
    {
        "번호": 19,
        "제목": "WCAG 2.2 (Web Content Accessibility Guidelines)",
        "저자/출처": "W3C",
        "URL": "https://www.w3.org/TR/WCAG22/",
        "발행일": "2023-10-05",
        "요약 (한국어)": "웹 접근성 표준 2.2. 터치 타깃 24×24px, 포커스 가시성, 드래그 대안, 인증 접근성 등 신규 기준. AA를 최소 준수 레벨로 권장.",
        "관련 섹션": "개발 주의점 - 접근성",
    },
    {
        "번호": 20,
        "제목": "Front-End Security Best Practices",
        "저자/출처": "SecureFlag",
        "URL": "https://blog.secureflag.com/2025/11/17/front-end-security-best-practices/",
        "발행일": "2025-11-17",
        "요약 (한국어)": "프론트엔드 보안 실무 가이드. XSS 방어(CSP, 새니타이즈), 토큰 저장(HttpOnly 쿠키), 입력 검증, 의존성 관리 등 클라이언트 측 위협 대응.",
        "관련 섹션": "개발 주의점 - 보안",
    },
    {
        "번호": 21,
        "제목": "OWASP Top 10:2025",
        "저자/출처": "OWASP Foundation",
        "URL": "https://owasp.org/www-project-top-ten/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "웹 애플리케이션 10대 보안 위험. 2025년 기준 Broken Access Control(#1), Security Misconfiguration(#2), 공급망 실패(#3)가 최상위. XSS는 Injection 범주.",
        "관련 섹션": "개발 주의점 - 보안",
    },
    {
        "번호": 22,
        "제목": "Modern Frontend Security: Beyond XSS and CSRF in 2025",
        "저자/출처": "Capture The Bug",
        "URL": "https://capturethebug.xyz/Blogs/Modern-Frontend-Security-Protecting-Your-Application-Beyond-XSS-and-CSRF-in-2025",
        "발행일": "2025-01-01",
        "요약 (한국어)": "클라이언트로 로직이 이동하면서 커진 프론트엔드 공격면 분석. 클라이언트 로직 유출, CSP 설정, 토큰·시크릿 관리, 서드파티 스크립트 위험을 다룸.",
        "관련 섹션": "개발 주의점 - 보안",
    },
    {
        "번호": 23,
        "제목": "TanStack Query (서버 상태 관리)",
        "저자/출처": "TanStack (Tanner Linsley)",
        "URL": "https://tanstack.com/query/latest",
        "발행일": "2025-01-01",
        "요약 (한국어)": "서버 상태(캐싱·동기화·무효화)를 전담하는 라이브러리. 서버 상태와 클라이언트 UI 상태를 분리하라는 현대 상태 관리 원칙의 대표 도구.",
        "관련 섹션": "개발 주의점 - 상태 관리",
    },
    {
        "번호": 24,
        "제목": "ESLint 9 Flat Config Tutorial",
        "저자/출처": "DEV Community",
        "URL": "https://dev.to/aolyang/eslint-9-flat-config-tutorial-2bm5",
        "발행일": "2025-01-01",
        "요약 (한국어)": "ESLint 9 flat config(eslint.config.js) 설정 튜토리얼. 기존 .eslintrc와의 차이, TypeScript·React 플러그인 구성, Prettier와의 역할 분담.",
        "관련 섹션": "코드 가이드 - 린팅",
    },
    {
        "번호": 25,
        "제목": "Configuring Prettier & ESLint for TypeScript/JavaScript",
        "저자/출처": "TeachMeIDEA",
        "URL": "https://teachmeidea.com/configuring-prettier-and-eslint-for-typescript-javascript-projects/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Prettier(포매팅)와 ESLint(품질)의 역할 분리, eslint-config-prettier로 충돌 제거하는 표준 설정. Prettier는 버그를 잡지 않고 일관 포매팅만 담당.",
        "관련 섹션": "코드 가이드 - 포매팅",
    },
    {
        "번호": 26,
        "제목": "Mastering ESLint Config — Feature-Sliced Design",
        "저자/출처": "Feature-Sliced Design",
        "URL": "https://feature-sliced.design/blog/mastering-eslint-config",
        "발행일": "2025-01-01",
        "요약 (한국어)": "FSD(계층·슬라이스) 아키텍처와 ESLint로 계층 의존 규칙·public API 경계를 강제하는 방법. 대규모 프론트엔드의 폴더 구조와 결합 관리.",
        "관련 섹션": "코드 가이드 - 폴더 구조",
    },
    {
        "번호": 27,
        "제목": "typescript-eslint",
        "저자/출처": "typescript-eslint Team",
        "URL": "https://typescript-eslint.io/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "TypeScript용 ESLint 통합. flat config 지원, recommendedTypeChecked 등 타입 인지 규칙, consistent-type-imports·no-explicit-any 등 품질 규칙.",
        "관련 섹션": "코드 가이드 - 린팅",
    },
    {
        "번호": 28,
        "제목": "Conventional Commits 1.0.0",
        "저자/출처": "Conventional Commits",
        "URL": "https://www.conventionalcommits.org/",
        "발행일": "2023-01-01",
        "요약 (한국어)": "커밋 메시지 컨벤션 명세. feat/fix/docs/refactor/chore 등 타입 접두로 자동 changelog·시맨틱 버저닝을 지원하는 협업 규약.",
        "관련 섹션": "코드 가이드 - Git",
    },
    {
        "번호": 29,
        "제목": "Comparing Angular, React, Vue, and Svelte: What You Need to Know",
        "저자/출처": "InfoWorld",
        "URL": "https://www.infoworld.com/article/3962039/what-you-need-to-know-about-angular-react-vue-and-svelte-popular-javascript-frameworks-compared.html",
        "발행일": "2025-01-01",
        "요약 (한국어)": "4대 프레임워크의 설계 철학·적합 사용처 비교. Angular(엔터프라이즈), React(생태계), Vue(균형), Svelte(성능)의 트레이드오프 정리.",
        "관련 섹션": "프레임워크별 상세 비교 / 선택 가이드",
    },
    {
        "번호": 30,
        "제목": "Svelte vs React in 2026: Performance & DX Compared",
        "저자/출처": "Strapi",
        "URL": "https://strapi.io/blog/svelte-vs-react-comparison",
        "발행일": "2026-01-01",
        "요약 (한국어)": "Svelte와 React의 성능·개발자 경험(DX) 비교. 번들 크기, 러닝커브, 생태계·채용 풀의 현실적 트레이드오프를 실무 관점에서 분석.",
        "관련 섹션": "선택 가이드",
    },
]


def create_xlsx(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 스타일 정의
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Malgun Gothic", size=10)
    link_font = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 행
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 45, 25, 50, 12, 60, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, ref in enumerate(REFERENCES, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약 (한국어)"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            ws.row_dimensions[row_idx].height = 60

            if col_idx == 1:  # 번호
                cell.alignment = center_align
                cell.font = Font(name="Malgun Gothic", size=10, bold=True)
            elif col_idx == 4:  # URL — 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = link_font
                cell.alignment = left_align
            elif col_idx == 5:  # 발행일
                cell.alignment = center_align
                cell.font = body_font
            else:
                cell.alignment = left_align
                cell.font = body_font

    # 틀 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "1F4E79"

    wb.save(output_path)
    print(f"XLSX 파일 생성 완료: {output_path}")
    print(f"총 참고문헌 수: {len(REFERENCES)}개")


if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "fe-framework-guide-references.xlsx")
    create_xlsx(output_path)
