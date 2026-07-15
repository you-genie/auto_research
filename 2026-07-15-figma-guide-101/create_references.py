"""
참고문헌 XLSX 파일 생성 스크립트
피그마 가이드 101 — 온보딩 클래스 (2025~2026)
"""

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

REFERENCES = [
    {
        "번호": 1,
        "제목": "Course overview: Figma Design for beginners (2025)",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/30848209492887-Course-overview-Figma-Design-for-beginners-2025",
        "발행일": "2025-01-01",
        "요약 (한국어)": "피그마 공식 초보자 코스 2025 개요. 인터페이스·프레임·도형·텍스트 기초부터 Auto Layout·컴포넌트·프로토타이핑까지의 학습 경로를 제시.",
        "관련 섹션": "전체 커리큘럼",
    },
    {
        "번호": 2,
        "제목": "Figma Design for beginners (섹션)",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/sections/30880632542743-Figma-Design-for-beginners",
        "발행일": "2025-01-01",
        "요약 (한국어)": "피그마 공식 초보자 학습 섹션. 기초 이후 컴포넌트와 Auto Layout으로 넘어가 '화면을 그린다'에서 '재사용 UI를 만든다'로 도약하도록 안내.",
        "관련 섹션": "학습 로드맵",
    },
    {
        "번호": 3,
        "제목": "Guide to auto layout",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/360040451373-Guide-to-auto-layout",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Auto Layout 공식 가이드. CSS Flexbox와 유사한 유연 레이아웃 시스템으로, 내용에 따라 요소가 자동 재배치되는 원리와 Hug/Fill/Fixed 설정을 설명.",
        "관련 섹션": "Auto Layout",
    },
    {
        "번호": 4,
        "제목": "Guide to variables in Figma",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/15339657135383-Guide-to-variables-in-Figma",
        "발행일": "2025-01-01",
        "요약 (한국어)": "변수(Variables) 공식 가이드. Color/Number/String/Boolean 타입과 Alias(참조)로 재사용 값을 저장하고 디자인 토큰을 구현하는 방법.",
        "관련 섹션": "스타일·변수·디자인 토큰",
    },
    {
        "번호": 5,
        "제목": "Modes for variables",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/15343816063383-Modes-for-variables",
        "발행일": "2025-01-01",
        "요약 (한국어)": "변수 모드(Modes) 공식 문서. 라이트/다크·브랜드·반응형 등 컨텍스트를 프레임 복제 없이 모드 전환으로 처리하는 테마링 방법.",
        "관련 섹션": "변수 - 모드(테마)",
    },
    {
        "번호": 6,
        "제목": "Overview of variables, collections, and modes",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/14506821864087-Overview-of-variables-collections-and-modes",
        "발행일": "2025-01-01",
        "요약 (한국어)": "변수·컬렉션·모드의 관계를 개괄. 컬렉션으로 변수를 묶고 모드로 값 세트를 전환하는 구조와 별칭(Alias) 참조 원리.",
        "관련 섹션": "변수 - 컬렉션/모드",
    },
    {
        "번호": 7,
        "제목": "Update 1: Tokens, variables, and styles",
        "저자/출처": "Figma Learn (Help Center)",
        "URL": "https://help.figma.com/hc/en-us/articles/18490793776023-Update-1-Tokens-variables-and-styles",
        "발행일": "2024-01-01",
        "요약 (한국어)": "디자인 토큰·변수·스타일의 차이와 관계 정리. 색·간격·타이포 등 디자인 속성을 토큰으로 관리하는 개념과 스타일 대비 변수의 장점.",
        "관련 섹션": "스타일 vs 변수",
    },
    {
        "번호": 8,
        "제목": "Config 2024 In Review",
        "저자/출처": "Figma Blog",
        "URL": "https://www.figma.com/blog/config-2024-recap/",
        "발행일": "2024-06-26",
        "요약 (한국어)": "Config 2024 요약. UI3 인터페이스 개편, Figma AI, Figma Slides, Dev Mode 강화, Code Connect(디자인 시스템 코드 연동) 발표.",
        "관련 섹션": "2024~2025 신기능",
    },
    {
        "번호": 9,
        "제목": "Config 2025: Pushing Design Further",
        "저자/출처": "Figma Blog",
        "URL": "https://www.figma.com/blog/config-2025-recap/",
        "발행일": "2025-05-07",
        "요약 (한국어)": "Config 2025 요약. Figma Make(프롬프트→앱), Figma Sites(웹 발행), Check designs 린터(변수 제안), 고급 드로잉, 새 Grid 시스템 발표.",
        "관련 섹션": "2024~2025 신기능",
    },
    {
        "번호": 10,
        "제목": "What Is Figma? Complete Guide to Features, Prototyping, Dev Mode",
        "저자/출처": "ALM Corp",
        "URL": "https://almcorp.com/blog/what-is-figma-complete-guide/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "피그마 종합 가이드. 컴포넌트·배리언트로 버튼·폼·카드의 모든 상태를 한 컴포넌트에 정의하는 법, 프로토타이핑, Dev Mode, 베스트 프랙티스.",
        "관련 섹션": "컴포넌트·배리언트 / Dev Mode",
    },
    {
        "번호": 11,
        "제목": "Figma for prototyping: a complete guide",
        "저자/출처": "Midrocket",
        "URL": "https://midrocket.com/en/guides/figma-for-prototyping/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "프로토타이핑 완전 가이드. 트리거·액션·전환 애니메이션으로 프레임을 연결하는 법과 Smart Animate 등 인터랙션 구현 방법.",
        "관련 섹션": "프로토타이핑",
    },
    {
        "번호": 12,
        "제목": "Master Figma in 2026: 10 Essential Tips for Beginners",
        "저자/출처": "Analytics Insight",
        "URL": "https://www.analyticsinsight.net/career/10-essential-figma-tips-and-skills-for-new-designers-2026",
        "발행일": "2026-01-01",
        "요약 (한국어)": "신입 디자이너를 위한 2026 필수 팁 10가지. Auto Layout 조기 습득, 컴포넌트화, 레이어 네이밍, 단축키 등 온보딩 핵심 습관.",
        "관련 섹션": "온보딩 팁",
    },
    {
        "번호": 13,
        "제목": "Design System Mastery with Figma Variables: 2025/2026 Best-Practice Playbook",
        "저자/출처": "Design Systems Collective",
        "URL": "https://www.designsystemscollective.com/design-system-mastery-with-figma-variables-the-2025-2026-best-practice-playbook-da0500ca0e66",
        "발행일": "2025-01-01",
        "요약 (한국어)": "변수 기반 디자인 시스템 실무 플레이북. 100개 컴포넌트가 아니라 토큰부터 시작하고 거버넌스를 적용해 점진적으로 컴포넌트를 늘리라는 조언.",
        "관련 섹션": "변수·토큰 온보딩 전략",
    },
    {
        "번호": 14,
        "제목": "Auto Layout in Figma — Figma Handbook",
        "저자/출처": "Design+Code",
        "URL": "https://designcode.io/figma-handbook-auto-layout/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "Auto Layout 실습 중심 핸드북. 방향·간격·패딩·정렬과 Hug/Fill 설정을 예제로 설명하며 반응형 컴포넌트 제작 흐름을 다룸.",
        "관련 섹션": "Auto Layout 실습",
    },
    {
        "번호": 15,
        "제목": "Schema 2025: Design Systems For A New Era",
        "저자/출처": "Figma Blog",
        "URL": "https://www.figma.com/blog/schema-2025-design-systems-recap/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "디자인 시스템 전용 컨퍼런스 Schema 2025 요약. 변수·토큰 거버넌스, Code Connect, 디자인–코드 동기화 등 시스템 확장 전략.",
        "관련 섹션": "디자인 시스템 / Dev Mode",
    },
    {
        "번호": 16,
        "제목": "Figma product news and release notes",
        "저자/출처": "Figma",
        "URL": "https://www.figma.com/release-notes/",
        "발행일": "2026-01-01",
        "요약 (한국어)": "피그마 공식 릴리스 노트. 신기능·개선사항의 최신 목록. 요금제·베타 상태가 자주 바뀌므로 클래스 진행 전 최신 상태 확인용.",
        "관련 섹션": "신기능 확인",
    },
    {
        "번호": 17,
        "제목": "NEW Figma 2025: Getting started, Beginner to Pro Class",
        "저자/출처": "Skillshare (Christine Vallaure)",
        "URL": "https://www.skillshare.com/en/classes/new-figma-2025-getting-started-the-beginner-to-pro-class/1399003407",
        "발행일": "2025-01-01",
        "요약 (한국어)": "2025 UI3 기준 초보→프로 클래스. 인터페이스부터 컴포넌트·Auto Layout까지 온보딩 커리큘럼 구성의 참고 사례.",
        "관련 섹션": "커리큘럼 참고",
    },
    {
        "번호": 18,
        "제목": "The Absolute Beginners Guide to Figma Design in 2025",
        "저자/출처": "Medium (Free Figma Templates)",
        "URL": "https://medium.com/@freefigmatemplates/the-absolute-beginners-guide-to-figma-design-in-2025-d30254cba0e0",
        "발행일": "2025-01-01",
        "요약 (한국어)": "완전 초보자를 위한 2025 입문 가이드. 인터페이스·프레임·기본 도구·간단 프로토타입까지 첫 화면 제작 흐름을 단계별로 안내.",
        "관련 섹션": "인터페이스·기본 도구",
    },
    {
        "번호": 19,
        "제목": "Figma Community",
        "저자/출처": "Figma",
        "URL": "https://www.figma.com/community",
        "발행일": "2026-01-01",
        "요약 (한국어)": "무료 템플릿·UI 키트·플러그인 공유 커뮤니티. 온보딩 실습 시 기존 UI 키트를 리믹스해 학습 속도를 높이는 데 활용.",
        "관련 섹션": "학습 리소스",
    },
    {
        "번호": 20,
        "제목": "Design Tokens: How to Sync Design and Code in Figma",
        "저자/출처": "Figma Resource Library",
        "URL": "https://www.figma.com/resource-library/design-tokens/",
        "발행일": "2025-01-01",
        "요약 (한국어)": "디자인 토큰으로 디자인과 코드를 동기화하는 방법. 변수를 코드로 내보내 개발과 값 체계를 일치시키는 실무 워크플로우.",
        "관련 섹션": "변수·토큰 / Dev Mode",
    },
]


def create_xlsx(output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "참고문헌"

    # 스타일 정의
    header_fill = PatternFill(start_color="7A3FF2", end_color="7A3FF2", fill_type="solid")
    even_fill = PatternFill(start_color="ECE3FB", end_color="ECE3FB", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Malgun Gothic", size=10)
    link_font = Font(name="Malgun Gothic", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 헤더 행
    headers = ["번호", "제목", "저자/출처", "URL", "발행일", "요약 (한국어)", "관련 섹션"]
    col_widths = [6, 45, 25, 50, 12, 60, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, ref in enumerate(REFERENCES, 2):
        is_even = (row_idx % 2 == 0)
        row_fill = even_fill if is_even else odd_fill

        values = [
            ref["번호"],
            ref["제목"],
            ref["저자/출처"],
            ref["URL"],
            ref["발행일"],
            ref["요약 (한국어)"],
            ref["관련 섹션"],
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            ws.row_dimensions[row_idx].height = 60

            if col_idx == 1:  # 번호
                cell.alignment = center_align
                cell.font = Font(name="Malgun Gothic", size=10, bold=True)
            elif col_idx == 4:  # URL — 하이퍼링크
                cell.hyperlink = ref["URL"]
                cell.font = link_font
                cell.alignment = left_align
            elif col_idx == 5:  # 발행일
                cell.alignment = center_align
                cell.font = body_font
            else:
                cell.alignment = left_align
                cell.font = body_font

    # 틀 고정 및 자동 필터
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(REFERENCES) + 1}"

    # 시트 탭 색상
    ws.sheet_properties.tabColor = "7A3FF2"

    wb.save(output_path)
    print(f"XLSX 파일 생성 완료: {output_path}")
    print(f"총 참고문헌 수: {len(REFERENCES)}개")


if __name__ == "__main__":
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "figma-guide-101-references.xlsx")
    create_xlsx(output_path)
